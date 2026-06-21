# -*- coding: utf-8 -*-
"""
Wochenplan Scheduler (pipeline: OG leaders → OG non-leaders → FR → Meetings)

- OG leaders (LA) → assigned to ALL of their dedicated OGs when present
  (includes Laufen; H.W. Ott leads Neuro & Laufen; active days defined by layout.json cells)

- OG non-leaders (FAs & AAs) → rotations first, then balance; coverage flags:
    • WENIGER ALS 2FA for MSK/Neuro/Onko/Thorax/Abdomen if total FAs < 2
    • KEIN FA IN BH / KEIN FA IN LI only for MSK & Abdomen (suppressed if < 2 FA flagged)
    • KEIN AA for all OGs except those in og_rules.json (rotation_or_leader_only)

- FR (Frontarzt) AFTER OGs
    • FR-only absences overlay includes that day’s Laufen assignees
    • optional manual FR-only exclusions (e.g., {"Donnerstag": {"H.W. Ott"}})

- Meetings LAST; standardized 4-pool priority matcher; excludes Laufen assignees per-day;
  Medizin BH & LI on Monday colored red (text only); per-meeting rules as configured.

Keeps macros/formatting (keep_vba=True). Uses exact cell maps.
"""

from dataclasses import dataclass, field
from typing import Set, Dict, List, Tuple, Optional
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
import random, re, os, zipfile, json, xml.etree.ElementTree as ET
from pathlib import Path

# ---------------- Constants ----------------

BH = "BH"
LI = "LI"

AA = "AA"   # Assistenzarzt
OA = "OA"   # Oberarzt
LA = "LA"   # Leitender Arzt

WEEKDAYS = ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag"]

# OG special rules — loaded from og_rules.json and organgruppen.json (edit via Streamlit UI)
def _load_og_rules():
    _path = Path(__file__).parent / "og_rules.json"
    with open(_path, encoding="utf-8") as _f:
        _r = json.load(_f)
    
    # Load OG list from organgruppen.json
    _og_path = Path(__file__).parent / "organgruppen.json"
    global OG_LIST_NO_LAUFEN, OG_LIST
    if _og_path.exists():
        with open(_og_path, encoding="utf-8") as _og_f:
            _og_data = json.load(_og_f)
            OG_LIST_NO_LAUFEN = _og_data.get("organgruppen", [])
            OG_LIST = OG_LIST_NO_LAUFEN  # No special Laufen handling
    else:
        # Fallback to hardcoded list if file doesn't exist
        OG_LIST_NO_LAUFEN = ["MSK", "Neuro", "Onko", "Thorax", "Abdomen", "Mammo", "Intervention/ Vaskulär", "Nuklearmedizin", "Laufen"]
        OG_LIST = OG_LIST_NO_LAUFEN
    
    global OG_PRIORITY_ORDER, USE_RANDOM_OG_SELECTION, OG_WEIGHTS_OA, OG_WEIGHTS_AA, OG_MAX_FAS, OG_MAX_AAS
    global SITE_COVERAGE_OVER_AVOID
    OG_PRIORITY_ORDER = _r.get("og_priority_order", OG_LIST_NO_LAUFEN)
    USE_RANDOM_OG_SELECTION = _r.get("use_random_og_selection", False)
    # When True, for OGs requiring both-site coverage a candidate from the missing
    # site is preferred over any same-site candidate even if they avoid the OG.
    # When False, the avoid/rotation preference outranks site coverage (legacy).
    SITE_COVERAGE_OVER_AVOID = _r.get("site_coverage_over_avoid", True)
    
    # Load separate OA and AA weights
    OG_WEIGHTS_OA = _r.get("og_weights_oa", _r.get("og_weights", {}))  # Fallback to old og_weights if needed
    OG_WEIGHTS_AA = _r.get("og_weights_aa", _r.get("og_weights", {}))
    
    OG_MAX_FAS = _r.get("og_max_fas", {})
    OG_MAX_AAS = _r.get("og_max_aas", {})
    
    # Set defaults for any missing OGs
    for og in OG_LIST:
        if og not in OG_WEIGHTS_OA:
            OG_WEIGHTS_OA[og] = 0.4 if og in ["Mammo", "Intervention/ Vaskulär"] else 0.6
        if og not in OG_WEIGHTS_AA:
            OG_WEIGHTS_AA[og] = 0.4 if og in ["Mammo", "Intervention/ Vaskulär"] else 0.6
        if og not in OG_MAX_FAS:
            OG_MAX_FAS[og] = None  # No limit
        if og not in OG_MAX_AAS:
            OG_MAX_AAS[og] = None  # No limit
    
    global US_VERTRETUNG_POOLS, OG_VERTRETUNG_OGS, OG_VERTRETUNG_POOLS
    global ROTATION_FIRST_AA_OGS
    US_VERTRETUNG_POOLS  = _r.get("us_vertretung_pools", {})
    OG_VERTRETUNG_OGS    = _r.get("og_vertretung_ogs", [])
    OG_VERTRETUNG_POOLS  = _r.get("og_vertretung_pools", {})
    ROTATION_FIRST_AA_OGS = _r.get("rotation_first_aa_ogs", [])

    return (
        set(_r.get("rotation_or_leader_only", [])),
        set(_r.get("warn_kein_aa", [])),
        set(_r.get("warn_weniger_als_2fa", [])),
        set(_r.get("warn_kein_fa_site", [])),
        set(_r.get("exclude_from_rapporte", [])),
    )

# OG Lists - will be populated from organgruppen.json
OG_LIST_NO_LAUFEN: List[str] = []
OG_LIST: List[str] = []

OG_PRIORITY_ORDER: List[str] = []
USE_RANDOM_OG_SELECTION: bool = False
SITE_COVERAGE_OVER_AVOID: bool = True
OG_WEIGHTS_OA: Dict[str, float] = {}
OG_WEIGHTS_AA: Dict[str, float] = {}
OG_MAX_FAS: Dict[str, Optional[int]] = {}
OG_MAX_AAS: Dict[str, Optional[int]] = {}
# US Vertretung: per-OG priority list of FAs for site-coverage substitution
# Key: OG name, Value: list of FA names in priority order
US_VERTRETUNG_POOLS: Dict[str, List[str]] = {}
# OG Vertretung: OGs that get a bracketed substitute when fully empty
OG_VERTRETUNG_OGS: List[str] = []
OG_VERTRETUNG_POOLS: Dict[str, List[str]] = {}
# OGs whose AAs with the matching rotation are seated (independent of OG load,
# cap overruled) before the general AA load-balancing round. AAs only.
ROTATION_FIRST_AA_OGS: List[str] = []
# NOTE: the defaults above must stay ABOVE this call. _load_og_rules() assigns
# all of them as module globals; declaring their defaults below the call would
# reset them to empty after the load.
OG_ROTATION_OR_LEADER_ONLY, OG_WARN_KEIN_AA, TARGET_OG_FOR_ONE_FA, TARGET_OG_FOR_KEIN_FA_SITE, OG_EXCLUDE_FROM_RAPPORTE = _load_og_rules()
OGS_SKIP_KEIN_AA = set(OG_LIST) - OG_WARN_KEIN_AA


def _load_fr_rules() -> tuple:
    _path = Path(__file__).parent / "fr_rules.json"
    if not _path.exists():
        return [], {}, set()
    with open(_path, encoding="utf-8") as f:
        _r = json.load(f)
    return (
        list(_r.get("exclude_names", [])),
        {k: set(v) for k, v in _r.get("exclude_if_day", {}).items()},
        set(_r.get("exclude_from_frontarzt", [])),
    )

FR_EXCLUDE_NAMES:        List[str]            = []
FR_EXCLUDE_IF_DAY:       Dict[str, Set[str]]  = {}
FR_EXCLUDE_FROM_OG:      Set[str]             = set()

def reload_fr_rules() -> None:
    """Reload FR rules from fr_rules.json — call after saving via UI."""
    global FR_EXCLUDE_NAMES, FR_EXCLUDE_IF_DAY, FR_EXCLUDE_FROM_OG
    FR_EXCLUDE_NAMES, FR_EXCLUDE_IF_DAY, FR_EXCLUDE_FROM_OG = _load_fr_rules()

# Initial load
FR_EXCLUDE_NAMES, FR_EXCLUDE_IF_DAY, FR_EXCLUDE_FROM_OG = _load_fr_rules()



# ---------------- Data model ----------------

@dataclass
class Staff:
    name: str
    role: str            # AA | OA | LA
    site: str            # BH | LI
    leads_ogs: Set[str] = field(default_factory=set)        # for LA only
    rotations: Set[str] = field(default_factory=set)        # AA + FA(non-leader)
    avoid_ogs: Set[str] = field(default_factory=set)        # soft-avoid: only assigned here as last resort
    fr_excluded: bool = False                               # if True → never Frontarzt on any day
    fr_excluded_days: Set[str] = field(default_factory=set) # excluded on specific weekdays only
    is_cover: bool = False                                  # if True → Stellvertreter (not shown in absence list)

    # Counters
    meetings_count: int = 0  # Deprecated - kept for backwards compatibility
    meetings_count_week: int = 0
    meetings_count_montag: int = 0
    meetings_count_dienstag: int = 0
    meetings_count_mittwoch: int = 0
    meetings_count_donnerstag: int = 0
    meetings_count_freitag: int = 0
    fr_shifts_count: int = 0
    og_nonleader_count: int = 0
    aa_og_count: int = 0

    @property
    def is_fa(self) -> bool:
        return self.role in {OA, LA}

staff_by_name: Dict[str, Staff] = {}

def add_staff(name: str, role: str, site: str,
              leads_for: Optional[List[str]] = None,
              rotation: Optional[List[str]] = None,
              avoid: Optional[List[str]] = None,
              fr_excluded: bool = False,
              fr_excluded_days: Optional[List[str]] = None,
              is_cover: bool = False):
    leads = set(leads_for or []) if role == LA else set()
    rots  = set(rotation or [])
    avoids = set(avoid or [])
    staff_by_name[name] = Staff(
        name=name,
        role=role,
        site=site,
        leads_ogs=leads,
        rotations=rots,
        avoid_ogs=avoids,
        fr_excluded=fr_excluded,
        fr_excluded_days=set(fr_excluded_days or []),
        is_cover=is_cover,
    )

def rebuild_quick_views() -> None:
    """Rebuild the module-level quick-view lists from staff_by_name.
    Must be called after any mutation of staff_by_name."""
    global aa_bh, aa_li, oa_bh, oa_li
    global fa_all_bh, fa_all_li, la_bh, la_li, leaders_by_og
    aa_bh            = [s.name for s in staff_by_name.values() if s.role == AA and s.site == BH]
    aa_li            = [s.name for s in staff_by_name.values() if s.role == AA and s.site == LI]
    oa_bh            = [s.name for s in staff_by_name.values() if s.role == OA and s.site == BH]
    oa_li            = [s.name for s in staff_by_name.values() if s.role == OA and s.site == LI]
    fa_all_bh        = [s.name for s in staff_by_name.values() if s.is_fa and s.site == BH]
    fa_all_li        = [s.name for s in staff_by_name.values() if s.is_fa and s.site == LI]
    la_bh            = [s.name for s in staff_by_name.values() if s.role == LA and s.site == BH]
    la_li            = [s.name for s in staff_by_name.values() if s.role == LA and s.site == LI]
    leaders_by_og    = {
        og: sorted([s.name for s in staff_by_name.values() if og in s.leads_ogs])
        for og in OG_LIST
    }


def load_staff_from_json(path: str) -> None:
    """Load staff from a JSON file, replacing the current staff_by_name contents.
    JSON format: list of objects with keys name, role, site, leads_ogs, rotations, fr_excluded, is_cover.
    Also loads site_rules from the same JSON file.
    Rebuilds all quick-view lists automatically."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    
    # Load site_rules if present
    global SITE_RULES
    if isinstance(data, dict) and "site_rules" in data:
        SITE_RULES = data.get("site_rules", {})
        records = data.get("staff", [])
    else:
        # Old format: just array of staff
        SITE_RULES = {"BH": {"no_oa_vormittag": False}, "LI": {"no_oa_vormittag": True}}
        records = data
    
    staff_by_name.clear()
    for r in records:
        add_staff(
            name=r["name"],
            role=r["role"],
            site=r["site"],
            leads_for=r.get("leads_ogs", []),
            rotation=r.get("rotations", []),
            avoid=r.get("avoid_ogs", []),
            fr_excluded=r.get("fr_excluded", False),
            fr_excluded_days=r.get("fr_excluded_days", []),
            is_cover=r.get("is_cover", False),
        )
    rebuild_quick_views()


# Quick-view placeholders — populated by load_staff_from_json below
aa_bh = aa_li = oa_bh = oa_li = []
fa_all_bh = fa_all_li = la_bh = la_li = []
leaders_by_og: Dict[str, List[str]] = {}

# Site rules loaded from staff.json
SITE_RULES: Dict[str, Dict[str, bool]] = {}

# staff.json is the single source of truth — always load from it.
_staff_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "staff.json")
if not os.path.exists(_staff_json):
    raise FileNotFoundError(
        f"staff.json not found at {_staff_json}. "
        "Create it before running the scheduler."
    )
load_staff_from_json(_staff_json)

# ---------------- Sheet utils & ranges ----------------

# Populated by load_layout_from_json() at startup — do not edit here.
ABW_RANGES:        Dict[str, str]                                      = {}
NACHT_RANGES:      Dict[str, str]                                      = {}
SPAETDIENST_CELLS: Dict[str, Dict[str, str]]                           = {}
FR_CELLS:          Dict[str, Dict[str, Tuple[str, ...]]]               = {}
OG_CELLS:          Dict[str, Dict[str, Tuple[str, ...]]]               = {}
MEETING_CELLS:     Dict[str, Dict[str, Dict[str, Tuple[str, ...]]]]    = {}
MEDIZIN_MONDAY_CELLS: Dict[str, str]                                   = {}
VORDERGRUNDDIENST_CELLS: Dict[str, str]                                = {}
HINTERGRUNDDIENST_CELLS: Dict[str, str]                                = {}
DATE_CELLS:        Dict[str, str]                                      = {}
WEEKDAY_DATE_CELLS: Dict[str, str]                                     = {}
FEIERTAGE:         Set[str]                                            = set()
FEIERTAGE_MERGE_CELLS: Dict[str, str]                                  = {}

# Populated by fill_dienste_from_csv() — keyed by the day the Hintergrund shift occurs
# (Mon–Fri for Nacht Mo-Fr, "Sonntag" for 24h Sa/So weekend shift).
# _filter_candidates() looks up D-1 to exclude the person from meetings the next morning.
HINTERGRUND_BY_DAY: Dict[str, str]                                     = {}

# Set by fill_dienste_from_csv() — e.g. "2026-KW21". Used by assign_meetings() to
# record stats for rapporte with statistik_führen=true. Empty string = no CSV loaded.
CURRENT_KW: str = ""

# Path to stats.json — cross-week assignment history for tracked rapporte.
STATS_JSON = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stats.json")


def load_layout_from_json(path: str) -> None:
    """Load all cell-map constants from layout.json, replacing current values."""
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    ABW_RANGES.clear()
    ABW_RANGES.update(data["abw_ranges"])

    NACHT_RANGES.clear()
    NACHT_RANGES.update(data["nacht_ranges"])

    SPAETDIENST_CELLS.clear()
    for site, day_map in data["spaetdienst_cells"].items():
        SPAETDIENST_CELLS[site] = dict(day_map)

    FR_CELLS.clear()
    for site, day_map in data["fr_cells"].items():
        FR_CELLS[site] = {day: tuple(cells) for day, cells in day_map.items()}

    OG_CELLS.clear()
    for og, day_map in data["og_cells"].items():
        OG_CELLS[og] = {day: tuple(cells) for day, cells in day_map.items()}

    MEETING_CELLS.clear()
    for site, mtg_map in data["meeting_cells"].items():
        MEETING_CELLS[site] = {
            mtg: {day: tuple(cells) for day, cells in day_map.items()}
            for mtg, day_map in mtg_map.items()
        }

    MEDIZIN_MONDAY_CELLS.clear()
    MEDIZIN_MONDAY_CELLS.update(data.get("medizin_monday_cells", {}))

    VORDERGRUNDDIENST_CELLS.clear()
    VORDERGRUNDDIENST_CELLS.update(data.get("vordergrunddienst_cells", {}))

    HINTERGRUNDDIENST_CELLS.clear()
    HINTERGRUNDDIENST_CELLS.update(data.get("hintergrunddienst_cells", {}))

    DATE_CELLS.clear()
    DATE_CELLS.update(data.get("date_cells", {}))

    WEEKDAY_DATE_CELLS.clear()
    WEEKDAY_DATE_CELLS.update(data.get("weekday_date_cells", {}))

    FEIERTAGE.clear()
    FEIERTAGE.update(data.get("feiertage", []))

    FEIERTAGE_MERGE_CELLS.clear()
    FEIERTAGE_MERGE_CELLS.update(data.get("feiertage_merge_cells", {}))


# layout.json holds all Excel cell-coordinate maps.
_layout_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "layout.json")
if not os.path.exists(_layout_json):
    raise FileNotFoundError(
        f"layout.json not found at {_layout_json}. "
        "Create it before running the scheduler."
    )
load_layout_from_json(_layout_json)


# meeting_pools.json holds the priority-pool definitions for every Rapport.
MEETING_POOLS: Dict[str, dict] = {}

def load_meeting_pools_from_json(path: str) -> None:
    """Load meeting pool definitions from JSON, replacing current MEETING_POOLS.
    Keys beginning with '_' are not rapporte — they hold settings (e.g. '_settings')
    and are read separately, not iterated as meetings."""
    global STATS_RESPECT_INWEEK
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    settings = data.get("_settings", {}) if isinstance(data.get("_settings"), dict) else {}
    STATS_RESPECT_INWEEK = bool(settings.get("stats_respect_inweek", True))
    MEETING_POOLS.clear()
    MEETING_POOLS.update(data)

# When True, _stats_fair_pick narrows by in-week load (day then week) before
# applying the cross-week ratio, so a person already assigned other rapporte this
# week is deprioritized. When False, it uses the cross-week ratio alone (legacy).
# Set from meeting_pools.json "_settings".stats_respect_inweek (default True).
STATS_RESPECT_INWEEK: bool = True

_pools_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "meeting_pools.json")
if not os.path.exists(_pools_json):
    raise FileNotFoundError(
        f"meeting_pools.json not found at {_pools_json}. "
        "Create it before running the scheduler."
    )
load_meeting_pools_from_json(_pools_json)


def reload_og_rules() -> None:
    """Reload OG special rules from og_rules.json and organgruppen.json — call after saving via UI."""
    global OG_ROTATION_OR_LEADER_ONLY, OG_WARN_KEIN_AA, TARGET_OG_FOR_ONE_FA, TARGET_OG_FOR_KEIN_FA_SITE, OGS_SKIP_KEIN_AA, OG_EXCLUDE_FROM_RAPPORTE
    OG_ROTATION_OR_LEADER_ONLY, OG_WARN_KEIN_AA, TARGET_OG_FOR_ONE_FA, TARGET_OG_FOR_KEIN_FA_SITE, OG_EXCLUDE_FROM_RAPPORTE = _load_og_rules()
    OGS_SKIP_KEIN_AA = set(OG_LIST) - OG_WARN_KEIN_AA
    # US_VERTRETUNG_POOLS, OG_VERTRETUNG_OGS, OG_VERTRETUNG_POOLS are updated inside _load_og_rules


# ---------------- Stats (cross-week fairness for tracked rapporte) ----------------

def load_stats() -> dict:
    """Load stats.json, returning empty dict if the file does not exist yet."""
    if not os.path.exists(STATS_JSON):
        return {}
    with open(STATS_JSON, encoding="utf-8") as f:
        return json.load(f)


def save_stats(data: dict) -> None:
    """Write stats.json atomically."""
    tmp = STATS_JSON + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp, STATS_JSON)


def write_stats_to_sheet(wb) -> None:
    """
    Write stats.json data to the 'Statistik' sheet of the workbook.

    Layout:
      Col A: Year (written once per year, on the KW01 row)
      Col B: KW number (integer)
      Col C+: One column per tracked rapport (statistik_führen=true),
              containing the assigned person's name or empty string.

    Rows are ordered KW01–KW53 per year, from earliest to most recent year
    found across all tracked rapporte in stats.json.
    """
    if "Statistik" not in wb.sheetnames:
        return

    ws = wb["Statistik"]

    # Load data
    stats = load_stats()
    pools = MEETING_POOLS

    # Collect tracked rapporte in the order they appear in MEETING_POOLS
    tracked_keys = [k for k, v in pools.items() if v.get("statistik_führen")]
    if not tracked_keys:
        return

    # Build a set of all (year, kw) tuples across all tracked rapporte
    all_kws: set = set()
    for key in tracked_keys:
        for name_data in stats.get(key, {}).values():
            for kw_str in name_data.get("history", []):
                # Format: "2026-KW21"
                try:
                    year, kw = kw_str.split("-KW")
                    all_kws.add((int(year), int(kw)))
                except ValueError:
                    pass

    if not all_kws:
        return

    # Sort chronologically
    sorted_kws = sorted(all_kws)

    # Build lookup: {rapport_key: {kw_str: name}}
    lookup: dict = {}
    for key in tracked_keys:
        lookup[key] = {}
        for name, name_data in stats.get(key, {}).items():
            for kw_str in name_data.get("history", []):
                lookup[key][kw_str] = name

    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Write header row
    header = ["Jahr", "KW"] + tracked_keys
    for col, val in enumerate(header, start=1):
        ws.cell(row=1, column=col, value=val)

    # Write data rows
    prev_year = None
    for row_idx, (year, kw) in enumerate(sorted_kws, start=2):
        kw_str = f"{year}-KW{kw:02d}"

        # Year column — only write when year changes
        if year != prev_year:
            ws.cell(row=row_idx, column=1, value=year)
            prev_year = year
        else:
            ws.cell(row=row_idx, column=1, value=None)

        # KW column
        ws.cell(row=row_idx, column=2, value=kw)

        # Rapport columns
        for col_idx, key in enumerate(tracked_keys, start=3):
            name = lookup[key].get(kw_str, None)
            ws.cell(row=row_idx, column=col_idx, value=name)


def _stats_fair_pick(
    meeting_key: str,
    candidates: List[str],
    kw_str: str,
    rng: random.Random,
    day: str,
    stats_weight: Dict[str, float],
    stats: Optional[dict] = None,
    persist: bool = True,
) -> Optional[str]:
    """
    Cross-week fair pick for rapporte with statistik_führen=true.

    When STATS_RESPECT_INWEEK is True (default), candidates are narrowed by
    in-week load first — fewest meetings TODAY (meetings_count_<day>), then
    fewest meetings THIS WEEK (meetings_count_week) — and the cross-week
    count/stats_weight ratio is applied only within that narrowed group. This
    mirrors _fair_pick_pool and prevents someone already busy this week from
    taking a tracked rapport while a less-loaded eligible person exists. If all
    candidates are equally loaded in-week, the narrowing is a no-op and the pure
    cross-week ratio decides — so the legacy outcome is preserved whenever there
    is no in-week difference to act on.

    When STATS_RESPECT_INWEEK is False, only the cross-week ratio is used.

    stats_weight per person comes from meeting_pools.json (default 1.0).
    Picks randomly among ties, then writes the assignment to stats.json and
    increments the in-week counters so later rapporte see this assignment.
    """
    if not candidates or not kw_str:
        return None

    # Use the run-shared stats dict when provided (so every pick in this run sees
    # the running counts); otherwise load from disk. In a test run the caller
    # passes persist=False, so selection still uses the statistics but nothing is
    # written back to stats.json.
    if stats is None:
        stats = load_stats()
    rapport_stats = stats.setdefault(meeting_key, {})

    # Cross-week ratio: assignment_count / stats_weight (lower = assign sooner).
    def ratio(name: str) -> float:
        count = rapport_stats.get(name, {}).get("count", 0)
        weight = stats_weight.get(name, 1.0)
        return count / weight if weight else float("inf")

    pool = list(candidates)
    if STATS_RESPECT_INWEEK:
        # 1) fewest meetings today
        day_counter_name = f"meetings_count_{day.lower()}"
        min_today = min(getattr(staff_by_name[n], day_counter_name, 0) for n in pool)
        pool = [n for n in pool if getattr(staff_by_name[n], day_counter_name, 0) == min_today]
        # 2) fewest meetings this week
        min_week = min(staff_by_name[n].meetings_count_week for n in pool)
        pool = [n for n in pool if staff_by_name[n].meetings_count_week == min_week]

    # 3) lowest cross-week ratio within the (possibly narrowed) pool
    min_ratio = min(ratio(n) for n in pool)
    bucket = [n for n in pool if ratio(n) == min_ratio]
    pick = rng.choice(bucket)

    # Update the running counts in memory so later rapporte this run see this
    # assignment. Persist to stats.json only outside test mode.
    entry = rapport_stats.setdefault(pick, {"count": 0, "history": []})
    entry["count"] += 1
    entry["history"].append(kw_str)
    if persist:
        save_stats(stats)

    # Increment in-week counters (so other rapporte this week see this assignment)
    staff_by_name[pick].meetings_count += 1
    staff_by_name[pick].meetings_count_week += 1
    day_counter_name = f"meetings_count_{day.lower()}"
    setattr(staff_by_name[pick], day_counter_name,
            getattr(staff_by_name[pick], day_counter_name, 0) + 1)

    return pick


def _clear_cells(ws: Worksheet, cells: Tuple[str, ...]):
    for a1 in cells:
        ws[a1].value = ""

def cleanup_blocks(ws: Worksheet, *, clear_fr=True, clear_og=True, clear_meetings=True):
    """
    Clears the configured cells for FR, OG, and/or Meetings.
    Formats/macros are preserved (caller is responsible for saving the workbook).
    """
    if clear_fr:
        for site in FR_CELLS:
            for day in WEEKDAYS:
                _clear_cells(ws, FR_CELLS[site][day])

    if clear_og:
        for og in OG_CELLS:
            for day in WEEKDAYS:
                _clear_cells(ws, OG_CELLS[og][day])

    if clear_meetings:
        for site in MEETING_CELLS:
            for mtg in MEETING_CELLS[site]:
                for day in WEEKDAYS:
                    _clear_cells(ws, MEETING_CELLS[site][mtg].get(day, ()))
    
def tokens_from_val(val: str) -> List[str]:
    if not isinstance(val, str) or not val.strip(): return []
    return [p.strip() for p in re.split(r",|;|/|\n|•", val) if p.strip()]

def read_spaetdienst_by_day(ws: Worksheet) -> Dict[str, Dict[str, Set[str]]]:
    """
    Returns {"BH": {weekday: set(names)}, "LI": {weekday: set(names)}} for Spätdienst,
    reading from fixed cells defined in SPAETDIENST_CELLS.
    """
    out: Dict[str, Dict[str, Set[str]]] = {
        "BH": {d: set() for d in WEEKDAYS},
        "LI": {d: set() for d in WEEKDAYS},
    }

    for site, day_map in SPAETDIENST_CELLS.items():
        for day, a1 in day_map.items():
            v = ws[a1].value
            if v:
                for t in tokens_from_val(str(v)):
                    out[site][day].add(t)

    return out

def _add_from_range(ws: Worksheet, a1: str, dest: set):
    for row in ws[a1]:
        for cell in row:
            if cell.value:
                for t in tokens_from_val(str(cell.value)):
                    dest.add(t)

def read_absences_by_day(ws: Worksheet) -> Dict[str, Set[str]]:
    """Read absences from ABW_RANGES and NACHT_RANGES cells."""
    absences = {d: set() for d in WEEKDAYS}
    for d in WEEKDAYS:
        _add_from_range(ws, ABW_RANGES[d], absences[d])
        _add_from_range(ws, NACHT_RANGES[d], absences[d])
    return absences

def remove_covers_from_absences_visual(ws: Worksheet) -> None:
    """
    Remove covers (Stellvertreter) from the visual absence list in ABW_RANGES.
    This is called AFTER the pipeline has finished, purely for display purposes.
    The pipeline itself uses the full absence list (including covers).
    """
    # Find all staff who are covers
    cover_names = {s.name for s in staff_by_name.values() if s.is_cover}
    
    if not cover_names:
        return  # Nothing to do
    
    # For each day's absence range
    for day in WEEKDAYS:
        if day not in ABW_RANGES:
            continue
        
        cell_range = ABW_RANGES[day]
        
        # Read current names from the range
        current_names = []
        for row in ws[cell_range]:
            for cell in row:
                val = cell.value
                if val and isinstance(val, str):
                    for name in tokens_from_val(val):
                        if name:
                            current_names.append(name)
        
        # Filter out covers
        filtered_names = [n for n in current_names if n not in cover_names]
        
        # Clear ALL cells in range
        for row in ws[cell_range]:
            for cell in row:
                cell.value = None
        
        # Write back filtered list (from top)
        cells_list = [cell for row in ws[cell_range] for cell in row]
        for i, name in enumerate(filtered_names):
            if i < len(cells_list):
                cells_list[i].value = name

def _make_font(old: Font, *, bold: bool, color: str) -> Font:
    """
    Return a new Font that copies every attribute from *old* but overrides
    bold and color.  color must be a fully-specified 8-char ARGB hex string
    (e.g. "FFFF0000" for opaque red, "FF000000" for opaque black) so that
    openpyxl does not silently prepend "00" and produce a transparent colour.
    """
    return Font(
        name=old.name, size=old.size, bold=bold, italic=old.italic,
        underline=old.underline, strike=old.strike, vertAlign=old.vertAlign,
        charset=old.charset, scheme=old.scheme, family=old.family,
        outline=old.outline, shadow=old.shadow, condense=old.condense,
        extend=old.extend, color=color,
    )

def set_bold_red(ws: Worksheet, a1: str):
    ws[a1].font = _make_font(ws[a1].font or Font(), bold=True,  color="FFFF0000")

def set_black_normal(ws: Worksheet, a1: str):
    ws[a1].font = _make_font(ws[a1].font or Font(), bold=False, color="FF000000")

def set_red(ws: Worksheet, a1: str):
    ws[a1].font = _make_font(ws[a1].font or Font(), bold=False, color="FFFF0000")
    
def reset_all_counters():
    """
    Reset all per-person and per-meeting counters so that each run of the
    pipeline starts from a clean state.
    """
    # Per-person counters
    for s in staff_by_name.values():
        s.meetings_count = 0
        s.meetings_count_week = 0
        s.meetings_count_montag = 0
        s.meetings_count_dienstag = 0
        s.meetings_count_mittwoch = 0
        s.meetings_count_donnerstag = 0
        s.meetings_count_freitag = 0
        s.fr_shifts_count = 0
        s.og_nonleader_count = 0
        s.aa_og_count = 0

    # Per-meeting/pool counters
    POOL_COUNTS.clear()

    
def print_weekly_stats():
    """
    Print a simple overview of Frontarztdienste (FR) and Rapporte (meetings)
    per person for this run.
    """
    print("\n=== Wochenstatistik: Frontarztdienste & Rapporte ===")
    print(f"{'Name':25} {'Frontarzt':>3} {'Rapporte':>9}")
    print("-" * 40)
    for s in sorted(staff_by_name.values(), key=lambda x: x.name):
        # You can show everyone, or only those with at least 1 of either:
        if s.fr_shifts_count == 0 and s.meetings_count == 0:
            continue
        print(f"{s.name:25} {s.fr_shifts_count:>3} {s.meetings_count:>9}")
    print("-" * 40)

    

# ---------------- FR (Frontarzt) ----------------

def pick_fa_for_fr_shift(day: str, fa_pool: list,
                          absences_by_day: dict,
                          avoid=None,
                          rng: random.Random = random,
                          fr_og_excluded: Optional[Set[str]] = None):
    avoid = set(avoid or [])
    present = []
    globally_excluded = set(FR_EXCLUDE_NAMES)
    day_excluded = FR_EXCLUDE_IF_DAY.get(day, set())

    for n in fa_pool:
        if n in avoid:
            continue
        if n in absences_by_day.get(day, set()):
            continue
        if n in globally_excluded:
            continue
        if n in day_excluded:
            continue
        if fr_og_excluded and n in fr_og_excluded:
            continue
        s = staff_by_name.get(n)
        if not s:
            continue
        if s.fr_excluded or day in s.fr_excluded_days:
            continue
        present.append(n)

    if not present:
        return None, "no eligible FR candidate"

    minc = min(staff_by_name[n].fr_shifts_count for n in present)
    bucket = [n for n in present if staff_by_name[n].fr_shifts_count == minc]
    choice = rng.choice(bucket)
    staff_by_name[choice].fr_shifts_count += 1
    return choice, None

def assign_fr_shifts_to_cells(
    ws: Worksheet,
    absences_orig: Dict[str, Set[str]],
    rng: random.Random,
    *,
    extra_exclusions: Optional[Dict[str, Set[str]]] = None,
) -> None:
    # Build absence dict — extra_exclusions merged in if provided
    abs_fr: Dict[str, Set[str]] = {d: set(absences_orig.get(d, set())) for d in WEEKDAYS}
    if extra_exclusions:
        for d in WEEKDAYS:
            abs_fr[d].update(extra_exclusions.get(d, set()))

    # Build per-day set of persons assigned to any OG excluded from Frontarzt
    fr_og_excluded_by_day: Dict[str, Set[str]] = {d: set() for d in WEEKDAYS}
    for og in FR_EXCLUDE_FROM_OG:
        if og not in OG_CELLS:
            continue
        for day in WEEKDAYS:
            for a1 in OG_CELLS[og].get(day, ()):
                v = ws[a1].value
                if isinstance(v, str) and v.strip():
                    fr_og_excluded_by_day[day].add(v.strip())

    for day in WEEKDAYS:
        if day in FEIERTAGE:
            continue  # Skip holidays

        fr_og_excl = fr_og_excluded_by_day.get(day)

        # BH: Process all cells
        used = set()
        bh_rules          = SITE_RULES.get("BH", {})
        bh_no_oa_vormittag = bh_rules.get("no_oa_vormittag", False)
        bh_excl_vorm      = set(bh_rules.get("fr_excluded_vormittag", []))
        bh_excl_nach      = set(bh_rules.get("fr_excluded_nachmittag", []))

        if bh_no_oa_vormittag and FR_CELLS["BH"][day]:
            # First cell (Vormittag): LA only, minus Vormittag exclusions
            top_cell = FR_CELLS["BH"][day][0]
            pool_vorm = [n for n in la_bh if n not in bh_excl_vorm]
            pick, _ = pick_fa_for_fr_shift(day, pool_vorm, abs_fr, used, rng, fr_og_excl)
            ws[top_cell].value = pick or ""
            if pick:
                used.add(pick)
            # Remaining cells (Nachmittag): all FA, minus Nachmittag exclusions
            pool_nach = [n for n in fa_all_bh if n not in bh_excl_nach]
            for a1 in FR_CELLS["BH"][day][1:]:
                pick, _ = pick_fa_for_fr_shift(day, pool_nach, abs_fr, used, rng, fr_og_excl)
                ws[a1].value = pick or ""
                if pick:
                    used.add(pick)
        else:
            # All cells: all FA split by position
            for idx, a1 in enumerate(FR_CELLS["BH"][day]):
                excl = bh_excl_vorm if idx == 0 else bh_excl_nach
                pool = [n for n in fa_all_bh if n not in excl]
                pick, _ = pick_fa_for_fr_shift(day, pool, abs_fr, used, rng, fr_og_excl)
                ws[a1].value = pick or ""
                if pick:
                    used.add(pick)

        # LI: Same logic
        used = set()
        li_rules          = SITE_RULES.get("LI", {})
        li_no_oa_vormittag = li_rules.get("no_oa_vormittag", True)
        li_excl_vorm      = set(li_rules.get("fr_excluded_vormittag", []))
        li_excl_nach      = set(li_rules.get("fr_excluded_nachmittag", []))

        if li_no_oa_vormittag and FR_CELLS["LI"][day]:
            # First cell (Vormittag): LA only, minus Vormittag exclusions
            top_cell = FR_CELLS["LI"][day][0]
            pool_vorm = [n for n in la_li if n not in li_excl_vorm]
            pick, _ = pick_fa_for_fr_shift(day, pool_vorm, abs_fr, used, rng, fr_og_excl)
            ws[top_cell].value = pick or ""
            if pick:
                used.add(pick)
            # Remaining cells (Nachmittag): all FA, minus Nachmittag exclusions
            pool_nach = [n for n in fa_all_li if n not in li_excl_nach]
            for a1 in FR_CELLS["LI"][day][1:]:
                pick, _ = pick_fa_for_fr_shift(day, pool_nach, abs_fr, used, rng, fr_og_excl)
                ws[a1].value = pick or ""
                if pick:
                    used.add(pick)
        else:
            # All cells: all FA split by position
            for idx, a1 in enumerate(FR_CELLS["LI"][day]):
                excl = li_excl_vorm if idx == 0 else li_excl_nach
                pool = [n for n in fa_all_li if n not in excl]
                pick, _ = pick_fa_for_fr_shift(day, pool, abs_fr, used, rng, fr_og_excl)
                ws[a1].value = pick or ""
                if pick:
                    used.add(pick)



# ---------------- OG assignment ----------------

FA_COUNTS: Dict[str, Dict[str,int]] = {d:{og:0 for og in OG_LIST} for d in WEEKDAYS}
AA_COUNTS: Dict[str, Dict[str,int]] = {d:{og:0 for og in OG_LIST} for d in WEEKDAYS}

def reset_og_counts():
    for d in WEEKDAYS:
        for og in OG_LIST: 
            FA_COUNTS[d][og]=0
            AA_COUNTS[d][og]=0

def _first_empty_cell(ws: Worksheet, cells: Tuple[str,...]) -> Optional[str]:
    for a1 in cells:
        v = ws[a1].value
        if v is None or (isinstance(v,str) and v.strip()==""): return a1
    return None

def _already_listed(ws: Worksheet, cells: Tuple[str,...], name: str) -> bool:
    for a1 in cells:
        v = ws[a1].value
        if isinstance(v,str) and v.strip()==name: return True
    return False

def _cell_of_name(ws: Worksheet, cells: Tuple[str,...], name: str) -> Optional[str]:
    """Return the A1 address of the cell holding exactly `name`, or None."""
    for a1 in cells:
        v = ws[a1].value
        if isinstance(v,str) and v.strip()==name: return a1
    return None

def _remove_and_compact(ws: Worksheet, cells: Tuple[str,...], name: str) -> bool:
    """Remove `name` from the OG cell range and pull all entries below it up by
    one so no gap is left. Returns True if the name was found and removed.

    Only the non-empty string entries are compacted; trailing cells are cleared.
    Flag strings (KEIN FA …, WENIGER …) are normally written after all people,
    so in practice this preserves their relative order too — but the swap runs
    before flags (Round 3), so the range contains only names at this point.
    """
    values = [ws[a1].value for a1 in cells]
    kept = []
    found = False
    for v in values:
        if isinstance(v, str) and v.strip() == name:
            found = True
            continue
        kept.append(v)
    if not found:
        return False
    for i, a1 in enumerate(cells):
        ws[a1].value = kept[i] if i < len(kept) else ""
    return True

def assign_la_to_ogs(ws: Worksheet, absences_by_day: Dict[str, Set[str]]) -> Dict[str, Dict[str, int]]:
    """
    Assign every present LA to all of their dedicated OGs (no prioritization).
    OGs without cells for a day are skipped automatically.
    """
    reset_og_counts()

    for day in WEEKDAYS:
        if day in FEIERTAGE:
            continue  # Skip holidays
        abs_today = absences_by_day.get(day, set())

        for og in OG_LIST:
            cells = OG_CELLS[og][day]
            if not cells:
                continue

            for leader in leaders_by_og.get(og, []):
                if leader in abs_today:
                    continue
                if _already_listed(ws, cells, leader):
                    continue
                slot = _first_empty_cell(ws, cells)
                if slot:
                    ws[slot].value = leader
                    FA_COUNTS[day][og] += 1

    return FA_COUNTS

def _names_in_cells(ws: Worksheet, cells: Tuple[str,...]) -> List[str]:
    out=[]
    for a1 in cells:
        v = ws[a1].value
        if isinstance(v,str) and v.strip(): out.append(v.strip())
    return out

def _assigned_on_prior_days(ws: Worksheet, og: str, day: str) -> Set[str]:
    """Names already assigned to `og` on earlier weekdays of the current plan.
    Powers 'sticky' assignments: prefer whoever covered this OG earlier in the
    same week so coverage stays continuous instead of reshuffling day to day."""
    names: Set[str] = set()
    if day not in WEEKDAYS:
        return names
    idx = WEEKDAYS.index(day)
    for d in WEEKDAYS[:idx]:
        for nm in _names_in_cells(ws, OG_CELLS.get(og, {}).get(d, [])):
            names.add(nm)
    return names

def _sticky_choice(candidates: list, prior_names: Set[str], rng: random.Random):
    """Among already equally-qualified `candidates`, prefer those who were
    assigned to the same OG on a prior weekday (random among them); otherwise
    random among all candidates. Never reorders across qualification tiers."""
    if not candidates:
        return None
    sticky = [n for n in candidates if n in prior_names]
    return rng.choice(sticky) if sticky else rng.choice(candidates)

def _has_fa_from_site(ws: Worksheet, cells: Tuple[str,...], site: str) -> bool:
    for nm in _names_in_cells(ws,cells):
        # A US-Vertretung substitute is written as "Name (US)"; strip the suffix
        # so it resolves to the underlying person and counts as real site
        # coverage (suppresses the KEIN FA IN <site> flag in Round 3).
        core = nm[:-5].strip() if nm.endswith(" (US)") else nm
        s = staff_by_name.get(core)
        if s and s.is_fa and s.site==site: return True
    return False

def _has_aa(ws: Worksheet, cells: Tuple[str,...]) -> bool:
    for nm in _names_in_cells(ws,cells):
        s = staff_by_name.get(nm)
        if s and s.role==AA: return True
    return False

def _place_in_og(ws: Worksheet, day: str, og: str, name: str, count_for_fa: bool):
    cells = OG_CELLS[og][day]
    if not cells: return False
    if _already_listed(ws,cells,name): return False
    slot = _first_empty_cell(ws,cells)
    if not slot: return False
    ws[slot].value = name
    s = staff_by_name.get(name)
    # Note: og_nonleader_count and aa_og_count are now incremented with weights
    # in assign_nonleaders_to_ogs, not here
    if count_for_fa and s and s.role==OA:
        FA_COUNTS[day][og] += 1
    if not count_for_fa and s and s.role==AA:
        AA_COUNTS[day][og] += 1
    return True

def assign_nonleaders_to_ogs(ws: Worksheet, absences_by_day: Dict[str,Set[str]], rng: random.Random) -> Dict[str,Dict[str,int]]:
    """
    Assign non-leader FAs and AAs to OGs using hybrid OG-centric approach with weights.
    
    Strategy:
    1. Reset daily counters for each person
    2. For OAs: While pool not empty, pick OG with lowest FA_COUNT, assign best matching OA
    3. For AAs: Same process with separate counter
    4. Respects OG_WEIGHTS (e.g., 0.4 for Mammo allows 2 assignments, 0.6 allows 1)
    """
    for day in WEEKDAYS:
        if day in FEIERTAGE:
            continue  # Skip holidays
        abs_today = absences_by_day.get(day, set())
        
        # ===== ROUND 1: OAs =====
        present_oas = [n for n in (oa_bh + oa_li) if n not in abs_today]
        
        # Reset daily counters
        for name in present_oas:
            staff_by_name[name].og_nonleader_count = 0
        
        # Pool: All OAs with counter that can still fit smallest weight
        min_weight = min(OG_WEIGHTS_OA.values()) if OG_WEIGHTS_OA else 0.6
        pool = set(present_oas)
        
        while pool:
            # 1. Find OGs with free slots.
            # Cap (OG_MAX_FAS) is bypassed for an OG if at least one person in the
            # pool has a rotation in that specific OG — they are only blocked by the
            # cap, not by any other OG being understaffed.
            # rotation_or_leader_only OGs (e.g. Mammo, Nuklearmedizin) are normally
            # filled only by their LA leader (handled upstream). They become available
            # in this fill round ONLY when someone in the pool has them as a rotation —
            # and the candidate guard below ensures only rotation-holders are placed
            # there, never a no-rotation/other-rotation spillover.
            available_ogs = [og for og in OG_LIST
                           if (og not in OG_ROTATION_OR_LEADER_ONLY
                               or any(og in staff_by_name[n].rotations for n in pool))
                           and _first_empty_cell(ws, OG_CELLS[og][day]) is not None
                           and (
                               OG_MAX_FAS.get(og) is None
                               or FA_COUNTS[day][og] < OG_MAX_FAS[og]
                               or any(og in staff_by_name[n].rotations for n in pool)
                           )]

            if not available_ogs:
                break  # No free slots

            # 2. For each OG: Find compatible persons (counter + weight <= 1.0 AND not already in this OG).
            # When the cap is reached, only rotation-matched candidates may be placed.
            og_candidates = {}
            for og in available_ogs:
                og_weight = OG_WEIGHTS_OA.get(og, 0.6)
                cap = OG_MAX_FAS.get(og)
                cap_reached = cap is not None and FA_COUNTS[day][og] >= cap
                compatible = [n for n in pool
                            if staff_by_name[n].og_nonleader_count + og_weight <= 1.0
                            and not _already_listed(ws, OG_CELLS[og][day], n)
                            and (not cap_reached or og in staff_by_name[n].rotations)
                            and (og not in OG_ROTATION_OR_LEADER_ONLY
                                 or og in staff_by_name[n].rotations)]
                if compatible:
                    og_candidates[og] = compatible

            if not og_candidates:
                break  # No compatible OG-person combinations
            
            # 3. Choose OG with lowest FA_COUNT among all eligible OGs.
            # Rotation matches no longer filter OG selection — they only
            # influence candidate selection within the chosen OG (step 4).
            eligible_ogs = list(og_candidates.keys())
            minv = min(FA_COUNTS[day][og] for og in eligible_ogs)
            bucket = [og for og in eligible_ogs if FA_COUNTS[day][og] == minv]

            # 4. OG-Priority or Random tiebreak
            if USE_RANDOM_OG_SELECTION:
                chosen_og = rng.choice(bucket)
            else:
                chosen_og = sorted(bucket, key=lambda x: OG_PRIORITY_ORDER.index(x) if x in OG_PRIORITY_ORDER else 999)[0]

            # 5. Choose person from compatible candidates.
            #
            # Base priority tiers: in_rotation > no_rotation > other_rotation >
            # avoider. An avoider (chosen_og in their avoid_ogs) is only picked
            # when every non-avoider tier is empty.
            #
            # For both-site-coverage OGs (warn_kein_fa_site) where exactly ONE
            # site is already covered, the relative precedence of "missing site"
            # vs "avoid flag" is controlled by SITE_COVERAGE_OVER_AVOID (see the
            # _group_order selection below). In both modes, rotation still orders
            # candidates *within* a site/avoider group via _pick_from_group.
            # When the OG does not require both-site coverage, or both/neither
            # sites are covered, opposite/same collapse and the base ordering
            # (in_rotation > no_rotation > other_rotation, non-avoider before
            # avoider) applies.
            candidates = og_candidates[chosen_og]

            # Determine the missing site for this OG, if site-balance applies.
            missing_site = None
            if chosen_og in TARGET_OG_FOR_KEIN_FA_SITE:
                sites_present = {
                    staff_by_name[n].site
                    for n in staff_by_name
                    if _already_listed(ws, OG_CELLS[chosen_og][day], n)
                }
                if len(sites_present) == 1:
                    covered_site = next(iter(sites_present))
                    missing_site = LI if covered_site == BH else BH

            # Partition: avoiders are pulled out into the lowest tier regardless
            # of rotation or site (soft-avoid; rotation in an avoided OG is moot).
            avoiders     = [n for n in candidates if chosen_og in staff_by_name[n].avoid_ogs]
            non_avoiders = [n for n in candidates if chosen_og not in staff_by_name[n].avoid_ogs]

            def _rot_tier(name: str) -> int:
                """0=in_rotation, 1=no_rotation, 2=other_rotation."""
                s = staff_by_name[name]
                if chosen_og in s.rotations:
                    return 0
                if not s.rotations:
                    return 1
                return 2

            def _other_rot_score_oa(name):
                rot_ogs = [og for og in staff_by_name[name].rotations if og != chosen_og]
                if not rot_ogs:
                    return 0
                return max(FA_COUNTS[day].get(og, 0) for og in rot_ogs)

            def _pick_from_group(group: list):
                """Pick one name from a same-site (or site-agnostic) group using
                the base rotation tiers: in > no > other. Within other_rotation,
                tiebreak by most-filled rotation OG. Returns None if empty."""
                if not group:
                    return None
                prior = _assigned_on_prior_days(ws, chosen_og, day)
                t_in    = [n for n in group if _rot_tier(n) == 0]
                t_no    = [n for n in group if _rot_tier(n) == 1]
                t_other = [n for n in group if _rot_tier(n) == 2]
                if t_in:
                    return _sticky_choice(t_in, prior, rng)
                if t_no:
                    return _sticky_choice(t_no, prior, rng)
                if t_other:
                    max_score = max(_other_rot_score_oa(n) for n in t_other)
                    best = [n for n in t_other if _other_rot_score_oa(n) == max_score]
                    return _sticky_choice(best, prior, rng)
                return None

            pick = None
            if missing_site is not None:
                # Site coverage dominates: try opposite-site non-avoiders first,
                # then same-site non-avoiders, then opposite-site avoiders, then
                # same-site avoiders.
                opp_non_av  = [n for n in non_avoiders if staff_by_name[n].site == missing_site]
                same_non_av = [n for n in non_avoiders if staff_by_name[n].site != missing_site]
                opp_av      = [n for n in avoiders if staff_by_name[n].site == missing_site]
                same_av     = [n for n in avoiders if staff_by_name[n].site != missing_site]
                # Group precedence depends on SITE_COVERAGE_OVER_AVOID:
                #  True  -> site coverage dominates the avoid flag:
                #           opp_non_av -> opp_av -> same_non_av -> same_av
                #  False -> avoid/rotation dominates site coverage (legacy):
                #           opp_non_av -> same_non_av -> opp_av -> same_av
                # Rotation sub-tiers inside each group are applied identically in
                # both modes by _pick_from_group.
                if SITE_COVERAGE_OVER_AVOID:
                    _group_order = (opp_non_av, opp_av, same_non_av, same_av)
                else:
                    _group_order = (opp_non_av, same_non_av, opp_av, same_av)
                for group in _group_order:
                    pick = _pick_from_group(group)
                    if pick is not None:
                        break
            else:
                # No site preference: base tiers, non-avoiders before avoiders.
                pick = _pick_from_group(non_avoiders)
                if pick is None:
                    pick = _pick_from_group(avoiders)

            if pick is None:
                break  # Should not happen

            # 6. Site-balance SWAP (only for warn_kein_fa_site OGs).
            # If the chosen OG needs the opposite site but no opposite-site
            # candidate exists in the pool, the normal logic above falls back
            # to a same-site pick — leaving the OG single-sided. Before
            # accepting that, try to swap in an opposite-site OA who is already
            # placed in another (non-site-sensitive) OG. If a swap succeeds we
            # restart the loop; the same-site fallback `pick` is NOT placed and
            # stays in the pool for a future iteration.
            # The swap is a site-coverage mechanism, so it only runs when site
            # coverage outranks avoid/rotation. In the legacy mode
            # (SITE_COVERAGE_OVER_AVOID == False) it is skipped entirely, so the
            # OG keeps its same-site pick and any missing-site coverage is left
            # to the manual US-Vertretung pool (Round 2.5) instead.
            if SITE_COVERAGE_OVER_AVOID and chosen_og in TARGET_OG_FOR_KEIN_FA_SITE:
                sites_present = {
                    staff_by_name[n].site
                    for n in staff_by_name
                    if _already_listed(ws, OG_CELLS[chosen_og][day], n)
                }
                # Only act when exactly one site is covered AND the pick we are
                # about to place is from that same (already-covered) site —
                # i.e. the fallback genuinely left the OG single-sided.
                if (len(sites_present) == 1
                        and staff_by_name[pick].site in sites_present):
                    covered_site = next(iter(sites_present))
                    chosen_weight = OG_WEIGHTS_OA.get(chosen_og, 0.6)

                    # Find opposite-site OAs already placed in another OG that
                    # is itself NOT site-sensitive (so we don't rob a balance-
                    # critical OG to fix this one), and who can absorb the
                    # weight change after the swap.
                    swap_options = []  # (name, source_og)
                    for src_og in OG_LIST:
                        if src_og == chosen_og:
                            continue
                        if src_og in TARGET_OG_FOR_KEIN_FA_SITE:
                            continue
                        src_weight = OG_WEIGHTS_OA.get(src_og, 0.6)
                        for nm in _names_in_cells(ws, OG_CELLS[src_og][day]):
                            s = staff_by_name.get(nm)
                            if not s or not s.is_fa or s.role != OA:
                                continue
                            if s.site == covered_site:
                                continue  # same site — no help
                            if _already_listed(ws, OG_CELLS[chosen_og][day], nm):
                                continue
                            # Weight after swap: drop source OG weight, add chosen OG weight
                            new_count = s.og_nonleader_count - src_weight + chosen_weight
                            if new_count <= 1.0:
                                swap_options.append((nm, src_og))

                    if swap_options:
                        # Prioritise swap candidate by the same tiers used for
                        # normal selection, relative to the CHOSEN (target) OG.
                        swap_in_rot = [(nm, so) for nm, so in swap_options
                                       if chosen_og in staff_by_name[nm].rotations]
                        swap_no_rot = [(nm, so) for nm, so in swap_options
                                       if not staff_by_name[nm].rotations]
                        swap_other  = [(nm, so) for nm, so in swap_options
                                       if staff_by_name[nm].rotations
                                       and chosen_og not in staff_by_name[nm].rotations]
                        if swap_in_rot:
                            swap_name, source_og = rng.choice(swap_in_rot)
                        elif swap_no_rot:
                            swap_name, source_og = rng.choice(swap_no_rot)
                        elif swap_other:
                            swap_name, source_og = rng.choice(swap_other)
                        else:
                            swap_name = source_og = None

                        if swap_name:
                            # Remove swap_name from its source OG, compacting the
                            # range so no gap remains for entries below it.
                            removed = _remove_and_compact(ws, OG_CELLS[source_og][day], swap_name)
                            if removed:
                                if staff_by_name[swap_name].role == OA:
                                    FA_COUNTS[day][source_og] -= 1
                                staff_by_name[swap_name].og_nonleader_count -= \
                                    OG_WEIGHTS_OA.get(source_og, 0.6)
                                # Place swap_name into the chosen (target) OG
                                _place_in_og(ws, day, chosen_og, swap_name, count_for_fa=True)
                                staff_by_name[swap_name].og_nonleader_count += chosen_weight
                                # Re-evaluate pool membership for swap_name
                                if staff_by_name[swap_name].og_nonleader_count + min_weight > 1.0:
                                    pool.discard(swap_name)
                                else:
                                    pool.add(swap_name)
                                # The same-site fallback `pick` is left unplaced
                                # and remains in the pool for a future iteration.
                                continue

            # 7. Place person in OG
            _place_in_og(ws, day, chosen_og, pick, count_for_fa=True)
            
            # 8. Update counter
            og_weight = OG_WEIGHTS_OA.get(chosen_og, 0.6)
            staff_by_name[pick].og_nonleader_count += og_weight
            
            # 9. Remove from pool if can't fit any more OGs
            if staff_by_name[pick].og_nonleader_count + min_weight > 1.0:
                pool.discard(pick)
        
        # ===== ROUND 2: AAs =====
        present_aas = [n for n in (aa_bh + aa_li) if n not in abs_today]
        
        # Reset daily counters
        for name in present_aas:
            staff_by_name[name].aa_og_count = 0
        
        pool = set(present_aas)

        # --- Rotation-first prerun (AAs only) ---
        # For OGs listed in rotation_first_aa_ogs, seat every PRESENT AA who has
        # that OG in their rotations BEFORE the general load-balancing loop, so a
        # rotation-AA is never spent as spillover in another OG first. This mirrors
        # how LA leaders are pre-seated into the OGs they lead. The AA cap
        # (OG_MAX_AAS) is intentionally overruled here — rotations win — but each
        # person's total weight is still capped at 1.0, and an OG that runs out of
        # cells stops accepting further rotation-AAs. OAs are unaffected.
        if ROTATION_FIRST_AA_OGS:
            prerun_ogs = [og for og in OG_PRIORITY_ORDER if og in ROTATION_FIRST_AA_OGS]
            for og in ROTATION_FIRST_AA_OGS:
                if og not in prerun_ogs:
                    prerun_ogs.append(og)
            for og in prerun_ogs:
                cells = OG_CELLS.get(og, {}).get(day, [])
                if not cells:
                    continue
                og_weight = OG_WEIGHTS_AA.get(og, 0.6)
                # Sticky: seat AAs who covered this OG earlier in the week first,
                # so a returning rotation-AA keeps the slot when cells are scarce.
                prior = _assigned_on_prior_days(ws, og, day)
                ordered = sorted(present_aas, key=lambda n: 0 if n in prior else 1)
                for name in ordered:
                    s = staff_by_name[name]
                    if og not in s.rotations:
                        continue
                    if s.aa_og_count + og_weight > 1.0:
                        continue
                    if _already_listed(ws, cells, name):
                        continue
                    if _first_empty_cell(ws, cells) is None:
                        break  # OG full — remaining rotation-AAs can't be seated
                    if _place_in_og(ws, day, og, name, count_for_fa=False):
                        s.aa_og_count += og_weight
            # Drop anyone who can no longer fit the smallest AA weight.
            _min_w_aa = min(OG_WEIGHTS_AA.values()) if OG_WEIGHTS_AA else 0.6
            for name in list(pool):
                if staff_by_name[name].aa_og_count + _min_w_aa > 1.0:
                    pool.discard(name)

        while pool:
            # Same logic as OAs, but using aa_og_count and checking max_aas.
            # Cap (OG_MAX_AAS) is bypassed for an OG if at least one person in the
            # pool has a rotation in that specific OG.
            # rotation_or_leader_only OGs (e.g. Mammo, Nuklearmedizin) become
            # available in this fill round ONLY when someone in the pool has them
            # as a rotation; the candidate guard below ensures only rotation-holders
            # are ever placed there (no no-rotation/other-rotation spillover).
            available_ogs = [og for og in OG_LIST
                           if (og not in OG_ROTATION_OR_LEADER_ONLY
                               or any(og in staff_by_name[n].rotations for n in pool))
                           and _first_empty_cell(ws, OG_CELLS[og][day]) is not None
                           and (
                               OG_MAX_AAS.get(og) is None
                               or AA_COUNTS[day][og] < OG_MAX_AAS[og]
                               or any(og in staff_by_name[n].rotations for n in pool)
                           )]

            if not available_ogs:
                break

            og_candidates = {}
            for og in available_ogs:
                og_weight = OG_WEIGHTS_AA.get(og, 0.6)
                cap = OG_MAX_AAS.get(og)
                cap_reached = cap is not None and AA_COUNTS[day][og] >= cap
                compatible = [n for n in pool
                            if staff_by_name[n].aa_og_count + og_weight <= 1.0
                            and not _already_listed(ws, OG_CELLS[og][day], n)
                            and (not cap_reached or og in staff_by_name[n].rotations)
                            and (og not in OG_ROTATION_OR_LEADER_ONLY
                                 or og in staff_by_name[n].rotations)]
                if compatible:
                    og_candidates[og] = compatible
            
            if not og_candidates:
                break
            
            # 3. Choose OG with lowest AA_COUNT among all eligible OGs.
            eligible_ogs = list(og_candidates.keys())
            minv = min(AA_COUNTS[day][og] for og in eligible_ogs)
            bucket = [og for og in eligible_ogs if AA_COUNTS[day][og] == minv]

            # 4. OG-Priority or Random tiebreak
            if USE_RANDOM_OG_SELECTION:
                chosen_og = rng.choice(bucket)
            else:
                chosen_og = sorted(bucket, key=lambda x: OG_PRIORITY_ORDER.index(x) if x in OG_PRIORITY_ORDER else 999)[0]

            # 5. Choose person from compatible candidates.
            # Priority: in_rotation > no_rotation > other_rotation.
            # Within other_rotation: prefer candidate whose rotation OG is
            # most filled (highest AA_COUNT) — least likely to need them there.
            candidates = og_candidates[chosen_og]

            in_rotation    = [n for n in candidates if chosen_og in staff_by_name[n].rotations]
            no_rotation    = [n for n in candidates if not staff_by_name[n].rotations]
            other_rotation = [n for n in candidates
                              if staff_by_name[n].rotations
                              and chosen_og not in staff_by_name[n].rotations]

            prior = _assigned_on_prior_days(ws, chosen_og, day)
            if in_rotation:
                pick = _sticky_choice(in_rotation, prior, rng)
            elif no_rotation:
                pick = _sticky_choice(no_rotation, prior, rng)
            elif other_rotation:
                # Tiebreak: prefer candidate whose rotation OG is most filled
                def _other_rot_score_aa(name):
                    rot_ogs = [og for og in staff_by_name[name].rotations if og != chosen_og]
                    if not rot_ogs:
                        return 0
                    return max(AA_COUNTS[day].get(og, 0) for og in rot_ogs)
                max_score = max(_other_rot_score_aa(n) for n in other_rotation)
                best = [n for n in other_rotation if _other_rot_score_aa(n) == max_score]
                pick = _sticky_choice(best, prior, rng)
            else:
                break
            
            _place_in_og(ws, day, chosen_og, pick, count_for_fa=False)
            
            og_weight = OG_WEIGHTS_AA.get(chosen_og, 0.6)
            staff_by_name[pick].aa_og_count += og_weight
            
            if staff_by_name[pick].aa_og_count + min_weight > 1.0:
                pool.discard(pick)
        
        # ===== ROUND 2.5: US Vertretung and OG Vertretung substitutions =====
        abs_today = absences_by_day.get(day, set())

        for og in OG_LIST:
            cells = OG_CELLS[og][day]

            # --- US Vertretung ---
            # Fires when OG has warn_kein_fa_site flag and a site is missing.
            # Finds first present FA from the missing site in the US priority list
            # and writes "Name (US)" into the next empty cell.
            if og in TARGET_OG_FOR_KEIN_FA_SITE:
                for missing_site in (BH, LI):
                    if not _has_fa_from_site(ws, cells, missing_site):
                        priority_list = US_VERTRETUNG_POOLS.get(og, [])
                        for candidate in priority_list:
                            if (candidate not in abs_today
                                    and candidate in staff_by_name
                                    and staff_by_name[candidate].site == missing_site):
                                slot = _first_empty_cell(ws, cells)
                                if slot:
                                    ws[slot].value = f"{candidate} (US)"
                                break  # One US per missing site

            # --- OG Vertretung ---
            # Fires when OG is in og_vertretung_ogs and has no FA or AA assigned at all.
            if og in OG_VERTRETUNG_OGS:
                # Check if OG is completely empty (no FA, no AA)
                has_any = any(
                    ws[c].value and str(ws[c].value).strip()
                    for c in cells
                    if c and ws[c].value
                )
                if not has_any:
                    priority_list = OG_VERTRETUNG_POOLS.get(og, [])
                    for candidate in priority_list:
                        if candidate not in abs_today and candidate in staff_by_name:
                            slot = _first_empty_cell(ws, cells)
                            if slot:
                                ws[slot].value = f"({candidate})"
                            break  # One vertretung per OG

        # ===== ROUND 3: Coverage flags =====
        for og in OG_LIST:
            cells = OG_CELLS[og][day]
            wrote_fa_shortage = False
            if og in TARGET_OG_FOR_ONE_FA and FA_COUNTS[day][og] < 2:
                slot = _first_empty_cell(ws,cells)
                if slot: ws[slot].value = "WENIGER ALS 2FA"; set_bold_red(ws,slot)
                wrote_fa_shortage = True
            if not wrote_fa_shortage and og in TARGET_OG_FOR_KEIN_FA_SITE:
                if not _has_fa_from_site(ws,cells,BH):
                    slot=_first_empty_cell(ws,cells)
                    if slot: ws[slot].value="KEIN FA IN BH"; set_bold_red(ws,slot)
                if not _has_fa_from_site(ws,cells,LI):
                    slot=_first_empty_cell(ws,cells)
                    if slot: ws[slot].value="KEIN FA IN LI"; set_bold_red(ws,slot)
            if og not in OGS_SKIP_KEIN_AA and not _has_aa(ws,cells):
                slot=_first_empty_cell(ws,cells)
                if slot: ws[slot].value="KEIN AA"; set_bold_red(ws,slot)
    
    return FA_COUNTS

# ---------------- Meetings (Rapporte): standardized 4-pool system ----------------

# Fairness counters for meeting pools
POOL_COUNTS = defaultdict(int)  # key: (meeting_key, pool_index, name) -> count

def _group_names(group: str, site: str) -> List[str]:
    if group == "AA":
        return aa_bh if site == BH else aa_li
    if group == "OA":
        return oa_bh if site == BH else oa_li
    if group == "LA":
        return [n for n in staff_by_name if staff_by_name[n].role == LA and staff_by_name[n].site == site]
    if group == "FA_ALL":
        return fa_all_bh if site == BH else fa_all_li
    raise ValueError(f"Unknown group: {group}")

def _filter_candidates(base: List[str], *, day: str, site: str,
                       absences: Dict[str, Set[str]],
                       spaetdienst_by_site_day: Dict[str, Dict[str, Set[str]]],
                       exclude_spaetdienst: Optional[str] = None,
                       exclude_names: Optional[Set[str]] = None,
                       exclude_if_day: Optional[Dict[str, Set[str]]] = None,
                       rapporte_excluded_names: Optional[Set[str]] = None,
                       exclude_hintergrund: bool = False) -> List[str]:
    c = [n for n in base if n not in absences.get(day, set())]
    if exclude_spaetdienst:
        c = [n for n in c if n not in spaetdienst_by_site_day.get(exclude_spaetdienst, {}).get(day, set())]
    if exclude_names:
        c = [n for n in c if n not in exclude_names]
    if exclude_if_day and day in exclude_if_day:
        c = [n for n in c if n not in exclude_if_day[day]]
    if rapporte_excluded_names:
        c = [n for n in c if n not in rapporte_excluded_names]
    if exclude_hintergrund:
        # Look up who had Hintergrund the previous night (D-1).
        # Only exclude if the person belongs to the same site as the pool.
        # Monday: the weekend shift may be split across Samstag and Sonntag,
        # so exclude both to be safe.
        if day == "Montag":
            for prev_key in ("Sonntag", "Samstag"):
                prev_hintergrund = HINTERGRUND_BY_DAY.get(prev_key)
                if prev_hintergrund:
                    s = staff_by_name.get(prev_hintergrund)
                    if s and s.site == site:
                        c = [n for n in c if n != prev_hintergrund]
        else:
            idx = WEEKDAYS.index(day)
            prev_key = WEEKDAYS[idx - 1]
            prev_hintergrund = HINTERGRUND_BY_DAY.get(prev_key)
            if prev_hintergrund:
                s = staff_by_name.get(prev_hintergrund)
                if s and s.site == site:
                    c = [n for n in c if n != prev_hintergrund]
    return c

def _fair_pick_pool(meeting_key: str, pool_idx: int, candidates: List[str], 
                    rng: random.Random, day: str) -> Optional[str]:
    """
    Fair pick for a given meeting pool with per-day and per-week balancing:
      1) Prefer people with FEWER meetings TODAY (meetings_count_<day>)
      2) Among those, prefer FEWER meetings THIS WEEK (meetings_count_week)
      3) Random choice among equals (using SEED for reproducibility)
    
    Note: POOL_COUNTS is no longer used - day and week counters provide better fairness.
    """
    if not candidates:
        return None

    # Step 1: Prefer people with fewer meetings today
    day_counter_name = f"meetings_count_{day.lower()}"
    min_today = min(getattr(staff_by_name[n], day_counter_name, 0) for n in candidates)
    cand_by_today = [n for n in candidates if getattr(staff_by_name[n], day_counter_name, 0) == min_today]

    # Step 2: Among those, prefer people with fewer meetings this week
    min_week = min(staff_by_name[n].meetings_count_week for n in cand_by_today)
    bucket = [n for n in cand_by_today if staff_by_name[n].meetings_count_week == min_week]

    # Step 3: Random choice with SEED
    return rng.choice(bucket)

def _bump_pool(meeting_key: str, pool_idx: int, name: str):
    POOL_COUNTS[(meeting_key, pool_idx, name)] += 1

def _assign(ws: Worksheet, a1: str, text: str, style: Optional[str]=None):
    ws[a1].value = text
    if style=="red_bold": set_bold_red(ws, a1)
    elif style=="black": set_black_normal(ws, a1)

def write_medizin_placeholders_monday(ws: Worksheet):
    for a1 in MEDIZIN_MONDAY_CELLS.values():
        if a1 and isinstance(a1, str) and a1.strip():
            _assign(ws, a1, "BITTE EINTRAGEN", "red_bold")
    
def assign_meeting_by_pools(
    ws: Worksheet,
    *,
    rng: random.Random,
    meeting_key: str,          # e.g. "BH|Medizin (07:45-08:00)"
    site: str,                 # BH or LI
    day: str,
    cells: Tuple[str, ...],
    pools: List[dict],         # up to 4 pools
    absences: Dict[str, Set[str]],
    spaetdienst: Dict[str, Dict[str, Set[str]]],
    rapporte_excluded_names: Set[str],
    monday_style: Optional[str] = None,       # e.g., "red"
    fallback_text: Optional[str] = "FÄLLT AUS",
    fallback_style: Optional[str] = "red_bold",
    statistik_führen: bool = False,
    stats_weight: Optional[Dict[str, float]] = None,
    stats: Optional[dict] = None,
    persist_stats: bool = True,
):
    for a1 in cells:
        placed = False
        for idx, pool in enumerate(pools, start=1):
            ptype = pool.get("type")
            style = pool.get("style")

            # Expand candidates for this pool
            if ptype == "names":
                base = list(pool.get("names", []))
            elif ptype == "group":
                base = _group_names(pool["group"], pool.get("site", site))
            elif ptype == "spaetdienst_aa":
                pool_site = pool.get("site", site)
                base = [n for n in spaetdienst[pool_site][day]
                        if staff_by_name.get(n) and staff_by_name[n].role == AA
                        and staff_by_name[n].site == pool_site]
                # Filter for absence, exclude_names, exclude_if_day only —
                # site is already guaranteed by the base list construction above.
                # Bypasses _fair_pick_pool: the Spätdienst AA is deterministic, not rotated.
                cands = _filter_candidates(
                    base, day=day, site=site, absences=absences,
                    spaetdienst_by_site_day=spaetdienst,
                    exclude_names=set(pool.get("exclude_names") or []) or None,
                    exclude_if_day={k: set(v) for k, v in pool.get("exclude_if_day", {}).items()} if pool.get("exclude_if_day") else None,
                    rapporte_excluded_names=rapporte_excluded_names,
                    exclude_hintergrund=False,
                )
                if cands:
                    pick = cands[0]
                    _assign(ws, a1, pick, style)
                    if monday_style and day == "Montag":
                        if monday_style == "red": set_red(ws, a1)
                    _bump_pool(meeting_key, idx, pick)
                    staff_by_name[pick].meetings_count += 1
                    staff_by_name[pick].meetings_count_week += 1
                    day_counter_name = f"meetings_count_{day.lower()}"
                    setattr(staff_by_name[pick], day_counter_name,
                            getattr(staff_by_name[pick], day_counter_name, 0) + 1)
                    placed = True
                    break
                continue  # No eligible Spätdienst AA — fall through to next pool
            elif ptype == "hintergrund_vortag":
                # Sunday-first priority: try each candidate individually through the
                # full filter so all exclusions apply; use the first one that survives.
                if day == "Montag":
                    ordered = [HINTERGRUND_BY_DAY.get("Sonntag"),
                               HINTERGRUND_BY_DAY.get("Samstag")]
                else:
                    prev_key = WEEKDAYS[WEEKDAYS.index(day) - 1]
                    ordered = [HINTERGRUND_BY_DAY.get(prev_key)]

                pick = None
                for candidate in ordered:
                    if not candidate:
                        continue
                    s = staff_by_name.get(candidate)
                    if not s or s.site != pool.get("site", site):
                        continue  # Wrong site — skip
                    cands = _filter_candidates(
                        [candidate], day=day, site=site, absences=absences,
                        spaetdienst_by_site_day=spaetdienst,
                        exclude_spaetdienst=pool.get("exclude_spaetdienst"),
                        exclude_names=set(pool.get("exclude_names") or []) or None,
                        exclude_if_day={k: set(v) for k, v in pool.get("exclude_if_day", {}).items()} if pool.get("exclude_if_day") else None,
                        rapporte_excluded_names=rapporte_excluded_names,
                        exclude_hintergrund=False,  # Would be circular
                    )
                    if cands:
                        pick = cands[0]
                        break

                if pick:
                    _assign(ws, a1, pick, style)
                    if monday_style and day == "Montag":
                        if monday_style == "red": set_red(ws, a1)
                    _bump_pool(meeting_key, idx, pick)
                    staff_by_name[pick].meetings_count += 1
                    staff_by_name[pick].meetings_count_week += 1
                    day_counter_name = f"meetings_count_{day.lower()}"
                    setattr(staff_by_name[pick], day_counter_name,
                            getattr(staff_by_name[pick], day_counter_name, 0) + 1)
                    placed = True
                    break
                continue  # No eligible Hintergrund person — fall through to next pool

            else:
                raise ValueError(f"Unknown pool type: {ptype}")

            # Filter candidates through all exclusion rules
            cands = _filter_candidates(
                base, day=day, site=site, absences=absences,
                spaetdienst_by_site_day=spaetdienst,
                exclude_spaetdienst=pool.get("exclude_spaetdienst"),
                exclude_names=set(pool.get("exclude_names") or []) or None,
                exclude_if_day={k:set(v) for k,v in pool.get("exclude_if_day", {}).items()} if pool.get("exclude_if_day") else None,
                rapporte_excluded_names=rapporte_excluded_names,
                exclude_hintergrund=pool.get("exclude_hintergrund", False),
            )

            # Pick: use cross-week stats ratio for tracked rapporte,
            # UNLESS the winning pool is hintergrund_vortag — those picks are
            # never recorded in stats (the person had no choice in being assigned).
            pool_is_hintergrund = pool.get("type") == "hintergrund_vortag"
            if statistik_führen and CURRENT_KW and not pool_is_hintergrund:
                pick = _stats_fair_pick(
                    meeting_key, cands, CURRENT_KW, rng, day,
                    stats_weight or {},
                    stats=stats, persist=persist_stats,
                )
                # _stats_fair_pick already increments meetings_count_* counters
                # and writes stats.json — nothing else needed here.
            else:
                pick = _fair_pick_pool(meeting_key, idx, cands, rng, day)
                if pick:
                    _bump_pool(meeting_key, idx, pick)
                    staff_by_name[pick].meetings_count += 1
                    staff_by_name[pick].meetings_count_week += 1
                    day_counter_name = f"meetings_count_{day.lower()}"
                    current_count = getattr(staff_by_name[pick], day_counter_name, 0)
                    setattr(staff_by_name[pick], day_counter_name, current_count + 1)

            if pick:
                _assign(ws, a1, pick, style)
                if monday_style and day == "Montag":
                    if monday_style == "red": set_red(ws, a1)
                placed = True
                break

        if not placed and fallback_text is not None:
            _assign(ws, a1, fallback_text, fallback_style)

def assign_meetings(ws: Worksheet, absences: Dict[str, Set[str]], rng: random.Random, skip_stats: bool = False):
    write_medizin_placeholders_monday(ws)

    spaet = read_spaetdienst_by_day(ws)

    # One shared stats dict for the whole run. Tracked rapporte are always picked
    # using the statistics (real and test runs alike); in a test run
    # (skip_stats=True) the running counts are kept in memory only and never
    # written back to stats.json.
    run_stats = load_stats()
    persist_stats = not skip_stats

    # Build per-day set of persons assigned to any OG that is excluded from rapporte.
    # Reads directly from the already-written OG cells in the sheet.
    rapporte_excluded_by_day: Dict[str, Set[str]] = {d: set() for d in WEEKDAYS}
    for og in OG_EXCLUDE_FROM_RAPPORTE:
        if og not in OG_CELLS:
            continue
        for day in WEEKDAYS:
            for a1 in OG_CELLS[og].get(day, ()):
                v = ws[a1].value
                if isinstance(v, str) and v.strip():
                    rapporte_excluded_by_day[day].add(v.strip())

    # Iterate over all meetings defined in meeting_pools.json, in file order.
    # Keys beginning with '_' are settings, not rapporte — skip them.
    for meeting_key, cfg in MEETING_POOLS.items():
        if meeting_key.startswith("_"):
            continue
        site = cfg["site"]
        # Derive the meeting name from the key (format: "SITE|Meeting Name")
        mtg_name = meeting_key.split("|", 1)[1]

        # Look up the cell map in MEETING_CELLS
        mtg_cells = MEETING_CELLS.get(site, {}).get(mtg_name, {})
        if not mtg_cells:
            continue

        pools              = cfg.get("pools", [])
        fallback_text      = cfg.get("fallback_text", "FÄLLT AUS")
        roter_fallback     = cfg.get("roter_fallback_text", True)
        fallback_style     = "red_bold" if roter_fallback else "black"
        statistik_führen   = cfg.get("statistik_führen", False)
        stats_weight       = cfg.get("stats_weight", {}) if statistik_führen else {}

        if statistik_führen and not CURRENT_KW:
            print(f"Warning: statistik_führen=true for '{meeting_key}' but CURRENT_KW is not set "
                  f"(no CSV loaded?). Stats will not be recorded this run.")

        for day, cells in mtg_cells.items():
            if day in FEIERTAGE:
                continue  # Skip holidays
            if not cells:
                continue

            assign_meeting_by_pools(
                ws, rng=rng, meeting_key=meeting_key, site=site, day=day, cells=cells,
                pools=pools, absences=absences, spaetdienst=spaet,
                rapporte_excluded_names=rapporte_excluded_by_day.get(day, set()),
                monday_style=None,
                fallback_text=fallback_text, fallback_style=fallback_style,
                statistik_führen=statistik_führen,
                stats_weight=stats_weight,
                stats=run_stats,
                persist_stats=persist_stats,
            )



# ---------------- XLSM post-save patch ----------------

def patch_xlsm(output_path: str, input_path: str) -> None:
    """
    Restore parts that openpyxl silently drops when saving a .xlsm file.

    openpyxl rewrites xl/worksheets/sheet1.xml and xl/styles.xml correctly,
    but it also:
      • Drops xl/drawings/drawing2.xml and its rels (strips the chart for sheet2)
      • Rewrites xl/worksheets/_rels/sheet2.xml.rels to point at drawing1.xml
        instead of drawing2.xml (wrong file, chart disappears)
      • Drops xl/printerSettings/printerSettings1.bin and its rels reference
      • Drops xl/sharedStrings.xml and xl/calcChain.xml
      • Changes [Content_Types].xml Default for .bin from printerSettings to vbaProject
      • Changes the VML relationship ID in sheet1.xml from "rId3" to "anysvml"

    Strategy: rebuild the output archive starting from the INPUT (which has the
    correct structure), then override only the parts openpyxl legitimately changed:
    sheet1.xml (cell data) and styles.xml (font/style definitions). Rebuild
    sheet1.xml.rels to include both the input's non-VML relationships and openpyxl's
    "anysvml" VML ID (which is what the new sheet1.xml references).
    """
    VML_TYPE = (
        "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"
    )
    RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

    def read_archive(path: str) -> dict:
        data = {}
        with zipfile.ZipFile(path, "r") as zf:
            for name in zf.namelist():
                data[name] = zf.read(name)
        return data

    inp = read_archive(input_path)
    out = read_archive(output_path)

    # Start from input (correct rels, drawings, Content_Types, printerSettings, etc.)
    patched = dict(inp)

    # Override the parts openpyxl correctly rewrites
    for key in ("xl/worksheets/sheet1.xml", "xl/styles.xml"):
        if key in out:
            patched[key] = out[key]

    # Rebuild sheet1.xml.rels: keep all input relationships except VML,
    # then add VML with the "anysvml" ID that openpyxl embedded in sheet1.xml.
    inp_s1_rels = ET.fromstring(inp["xl/worksheets/_rels/sheet1.xml.rels"])
    rel_parts = []
    for rel in inp_s1_rels:
        if rel.get("Type") != VML_TYPE:
            rel_parts.append(
                f'<Relationship Id="{rel.get("Id")}"'
                f' Type="{rel.get("Type")}"'
                f' Target="{rel.get("Target")}"/>'
            )
    rel_parts.append(
        f'<Relationship Id="anysvml"'
        f' Type="{VML_TYPE}"'
        f' Target="../drawings/vmlDrawing1.vml"/>'
    )
    merged_rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        f'<Relationships xmlns="{RELS_NS}">'
        + "".join(rel_parts)
        + "</Relationships>"
    )
    patched["xl/worksheets/_rels/sheet1.xml.rels"] = merged_rels.encode("utf-8")

    # Write the patched archive back to output_path
    tmp = output_path + ".tmp"
    with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for name, data in patched.items():
            zout.writestr(name, data)
    os.replace(tmp, output_path)


# ---------------- CLI pipeline (OG → OG non-leaders → FR → Meetings) ----------------
# Default: bash> python wochenplan_scheduler.py -i "KW XX_leer.xlsm" -o "KW XX.xlsm" --seed 1234
# Skip cleanup: append -nocleanup
# keep intermediates: append --keep

# ===========================================================================
# CSV Import Functions
# ===========================================================================

def match_csv_name_to_staff(csv_name: str) -> Optional[str]:
    """
    Match a CSV name to the exact staff.json name by surname, with initial fallback for duplicates.
    
    CSV format: "Nachname(n) Vorname(n)" - last part is always first name
    staff.json format: "Initial(s) Nachname" - last part is always surname
    
    Strategy:
    1. Take all parts from CSV except the last (those are surname parts)
    2. Split by hyphens to get all surname candidates
    3. Match against staff.json surnames (last part of staff name)
    4. If multiple matches, filter by first initial
    5. Return exact staff.json name
    
    Examples:
        "Iacobut Andreea-Emilia" → "A. Iacobut"
        "Slezenkovska-Stojanoska Tanja" → "T. Slezenkovska"
        "Berberich Bettina Katharina" → "B. Berberich"
        "Ott Hans-Werner" → "H.W. Ott"
    """
    csv_parts = csv_name.strip().split()
    if len(csv_parts) < 2:
        return None
    
    # All parts except last = surname parts
    surname_parts = csv_parts[:-1]
    # Last part = first name
    first_name = csv_parts[-1]
    
    # Get first initial from first name (handle hyphens)
    first_initial = first_name.split('-')[0][0].upper() if first_name else None
    
    # Split surnames by hyphens and collect all surname candidates
    surname_candidates = []
    for part in surname_parts:
        if '-' in part:
            surname_candidates.extend(part.split('-'))
        else:
            surname_candidates.append(part)
    
    # Normalize: lowercase for comparison
    surname_candidates_lower = [s.lower() for s in surname_candidates]
    
    # Find all staff with matching surname
    matches = []
    for staff_name in staff_by_name.keys():
        staff_parts = staff_name.split()
        if len(staff_parts) < 2:
            continue
        
        # Last part of staff name is the surname
        staff_surname = staff_parts[-1].lower()
        
        if staff_surname in surname_candidates_lower:
            matches.append(staff_name)
    
    # If single match, return it
    if len(matches) == 1:
        return matches[0]
    
    # If multiple matches, filter by first initial
    if len(matches) > 1 and first_initial:
        initial_matches = [m for m in matches if m.split()[0][0].upper() == first_initial]
        if len(initial_matches) == 1:
            return initial_matches[0]
        elif len(initial_matches) > 1:
            # Still ambiguous - return first and warn
            print(f"Warning: Ambiguous match for {csv_name}: {initial_matches}, using {initial_matches[0]}")
            return initial_matches[0]
    
    # No match or still ambiguous
    if matches:
        print(f"Warning: Multiple surname matches for {csv_name}: {matches}, using {matches[0]}")
        return matches[0]
    
    print(f"Warning: No staff match found for CSV name: {csv_name}")
    return None


def fill_dienste_from_csv(ws: Worksheet, csv_path: str) -> None:
    """
    Read dienste and absences from CSV and fill into Excel template.
    
    Args:
        ws: openpyxl worksheet to fill
        csv_path: Path to CSV file with dienste data
    """
    import pandas as pd
    from datetime import datetime
    
    # Try UTF-8 first (better special character support), fallback to ISO-8859-1
    try:
        df = pd.read_csv(csv_path, sep=';', encoding='utf-8')
    except UnicodeDecodeError:
        df = pd.read_csv(csv_path, sep=';', encoding='ISO-8859-1')
    
    # Parse dates
    df['Datum'] = pd.to_datetime(df['Datum'], format='%d.%m.%Y')
    
    # Find Monday of the week (first day)
    mondays = df[df['Datum'].dt.dayofweek == 0]['Datum']
    if len(mondays) == 0:
        raise ValueError("No Monday found in CSV - cannot determine week start")
    monday_date = mondays.iloc[0]

    # Write date to T20 (if date_cells exist in layout)
    if DATE_CELLS and "first_monday" in DATE_CELLS:
        ws[DATE_CELLS["first_monday"]].value = monday_date.strftime('%d/%m/%Y')

    # Calculate and write KW (ISO week number); also set CURRENT_KW for stats tracking.
    global CURRENT_KW
    iso = monday_date.isocalendar()
    CURRENT_KW = f"{iso[0]}-KW{iso[1]:02d}"
    if DATE_CELLS and "kw_number" in DATE_CELLS:
        ws[DATE_CELLS["kw_number"]].value = f"KW {iso[1]}"

    # Dienst type mapping — loaded from bezeichnungen.json
    _bez_path = Path(__file__).parent / "bezeichnungen.json"
    with open(_bez_path, encoding="utf-8") as _f:
        _bez = json.load(_f)
    ABSENZ_TYPES = set(_bez.get("absenz", []))
    SKIP_TYPES   = set(_bez.get("skip", []))   # EIR, PEPVIRTUELL etc. — ignore silently

    # Day name mapping (German date to layout key)
    day_names = {
        0: "Montag", 1: "Dienstag", 2: "Mittwoch",
        3: "Donnerstag", 4: "Freitag", 5: "Samstag", 6: "Sonntag"
    }

    # --- Pass 1: prior weekend only (Saturday and Sunday before Monday) ----------
    # Populate HINTERGRUND_BY_DAY from the preceding Sat/Sun so that Monday's
    # rapport exclusion logic (_filter_candidates exclude_hintergrund) works
    # correctly. Nothing from these two days is written to any Excel cell.
    HINTERGRUND_BY_DAY.clear()
    prior_weekend = df[df['Datum'] < monday_date]
    for _, row in prior_weekend.iterrows():
        dienst_type = row['Bezeichnung'].strip() if isinstance(row['Bezeichnung'], str) else row['Bezeichnung']
        if any(tok in dienst_type for tok in SKIP_TYPES):
            continue
        csv_name = row['Suchname']
        matched_name = match_csv_name_to_staff(csv_name)
        if not matched_name:
            continue
        day_name = day_names[row['Datum'].dayofweek]
        if "Pikett_24h_Sa/So" in dienst_type or "Pikett_Nacht_Mo-Fr" in dienst_type:
            HINTERGRUND_BY_DAY[day_name] = matched_name

    # --- Pass 2: current week (Monday through Sunday) — cell writes --------------
    # All normal processing: absences, Nacht, Spät, Vordergrund, Hintergrund.
    # Rows from the prior weekend are excluded so they never reach any Excel cell.
    df_week = df[df['Datum'] >= monday_date]

    # Collect data by day and type
    absences_by_day = {d: [] for d in day_names.values()}
    nacht_by_day = {}
    spaet_by_day = {"BH": {}, "LI": {}}
    vordergrund_by_day = {}
    hintergrund_by_day = {}

    for _, row in df_week.iterrows():
        dienst_type = row['Bezeichnung'].strip() if isinstance(row['Bezeichnung'], str) else row['Bezeichnung']
        csv_name = row['Suchname']
        date = row['Datum']
        day_name = day_names[date.dayofweek]

        # Convert name format - match to staff.json
        matched_name = match_csv_name_to_staff(csv_name)
        if not matched_name:
            print(f"Skipping unknown person from CSV: {csv_name}")
            continue
        abbrev_name = matched_name

        # Categorize dienst.
        # All categories now match by substring (token contained anywhere in the
        # Bezeichnung), consistent with the Nachtdienst/Spätdienst/Pikett branches
        # below. Catalog tokens (skip, absenz) must be long, distinctive strings —
        # short tokens could collide across categories. .strip() above still guards
        # against manual leading/trailing whitespace in the CSV.
        if any(tok in dienst_type for tok in SKIP_TYPES):
            continue  # EIR, PEPVIRTUELL etc. — ignore entirely

        elif any(tok in dienst_type for tok in ABSENZ_TYPES):
            absences_by_day[day_name].append(abbrev_name)

        elif "Nachtdienst" in dienst_type:
            nacht_by_day[day_name] = abbrev_name

        # --- Class AA ---
        # Spätdienst Mo-Fr: always goes to spaetdienst_cells (weekday cell).
        # Pikett_Vormittag_Sa/So: always goes to vordergrund_by_day — the
        # existing write logic handles both weekends (Vordergrund cell) and
        # holidays on weekdays (merge cell via FEIERTAGE_MERGE_CELLS).
        # Tagdienst Sa/So always appears paired with Pikett_Vormittag_Sa/So —
        # skipped to avoid double-writes.
        elif "Spätdienst" in dienst_type:
            site = "BH" if dienst_type.startswith("Bh-") else "LI"
            spaet_by_day[site][day_name] = abbrev_name

        elif "Pikett_Vormittag_Sa/So" in dienst_type:
            vordergrund_by_day[day_name] = abbrev_name

        elif "Tagdienst Sa/So" in dienst_type:
            pass  # Always paired with Pikett_Vormittag_Sa/So — handled above

        # --- Class FA: Hintergrund (weekdays, weekends, holidays) ---
        # Both dienst types populate hintergrund_by_day (Excel cell).
        # HINTERGRUND_BY_DAY is also updated for weekday entries (incl. holidays)
        # so that hintergrund_vortag lookups work the next day.
        # For Sa/So entries HINTERGRUND_BY_DAY is NOT updated here — Pass 1
        # already set it from the prior weekend.
        elif "Pikett_Nacht_Mo-Fr" in dienst_type or "Pikett_24h_Sa/So" in dienst_type:
            hintergrund_by_day[day_name] = abbrev_name
            if date.dayofweek < 5:  # Mon–Fri: weekday or holiday
                HINTERGRUND_BY_DAY[day_name] = abbrev_name

        # Other Pikett types — write to Hintergrund cell only
        elif "Pikett" in dienst_type:
            hintergrund_by_day[day_name] = abbrev_name

        # Tagdienst weekday — normal working day, no special cell
        elif "Tagdienst" in dienst_type:
            pass
    
    # Write to Excel
    from openpyxl.styles import PatternFill
    gray_fill = PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid")
    
    # Absences - write each name to separate row (skip on holidays)
    for day, names in absences_by_day.items():
        if day in FEIERTAGE:
            continue  # Skip absences on holidays
        if names and day in ABW_RANGES:
            cell_range = ABW_RANGES[day]
            # Parse range: "T94:AC108" → start_col="T", start_row=94
            from openpyxl.utils import column_index_from_string, get_column_letter
            
            start_cell = cell_range.split(':')[0]
            # Extract column letter and row number
            col_letter = ''.join(c for c in start_cell if c.isalpha())
            start_row = int(''.join(c for c in start_cell if c.isdigit()))
            
            # Write each name to a new row
            for idx, name in enumerate(sorted(set(names))):
                cell = f"{col_letter}{start_row + idx}"
                ws[cell].value = name
    
    # Nachtdienst - single cell per day (write on all days including holidays)
    for day, name in nacht_by_day.items():
        if day in NACHT_RANGES:
            cell_range = NACHT_RANGES[day]
            first_cell = cell_range.split(':')[0]
            ws[first_cell].value = name
    
    # Spätdienst (normal days)
    for site in ["BH", "LI"]:
        for day, name in spaet_by_day[site].items():
            if day in FEIERTAGE:
                continue  # Handled separately below
            cell = SPAETDIENST_CELLS.get(site, {}).get(day)
            if cell:
                ws[cell].value = name
    
    # Vordergrunddienst (holidays and weekends)
    for day, name in vordergrund_by_day.items():
        if day in FEIERTAGE:
            # On holidays: Merge cells using FEIERTAGE_MERGE_CELLS
            merge_range = FEIERTAGE_MERGE_CELLS.get(day)
            if merge_range:
                ws.merge_cells(merge_range)
                # Write to first cell of range
                first_cell = merge_range.split(':')[0]
                ws[first_cell].value = name
        else:
            # Normal weekend: Write to Vordergrunddienst cell
            cell = VORDERGRUNDDIENST_CELLS.get(day)
            if cell:
                ws[cell].value = name
    
    # Hintergrunddienst (write on all days including holidays)
    for day, name in hintergrund_by_day.items():
        cell = HINTERGRUNDDIENST_CELLS.get(day)
        if cell:
            ws[cell].value = name
    
    # Gray out FR, OG, Rapporte, and Absences on holidays
    for day in FEIERTAGE:
        # Gray out Absence range
        if day in ABW_RANGES:
            cell_range = ABW_RANGES[day]
            # Parse range to get all cells
            start_cell, end_cell = cell_range.split(':')
            
            # Extract column letters and row numbers
            start_col = ''.join(c for c in start_cell if c.isalpha())
            start_row = int(''.join(c for c in start_cell if c.isdigit()))
            end_col = ''.join(c for c in end_cell if c.isalpha())
            end_row = int(''.join(c for c in end_cell if c.isdigit()))
            
            # Gray out all cells in the range
            from openpyxl.utils import column_index_from_string
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)
            
            for col_idx in range(start_col_idx, end_col_idx + 1):
                for row in range(start_row, end_row + 1):
                    from openpyxl.utils import get_column_letter
                    cell = f"{get_column_letter(col_idx)}{row}"
                    ws[cell].fill = gray_fill
        
        # Gray out FR cells
        for site in ["BH", "LI"]:
            fr_cells = FR_CELLS.get(site, {}).get(day, [])
            for cell in fr_cells:
                ws[cell].fill = gray_fill
        
        # Gray out OG cells
        for og in OG_CELLS:
            og_cells = OG_CELLS[og].get(day, [])
            for cell in og_cells:
                ws[cell].fill = gray_fill
        
        # Gray out Rapporte cells
        for site in ["BH", "LI"]:
            for mtg_name, mtg_days in MEETING_CELLS.get(site, {}).items():
                mtg_cells = mtg_days.get(day, [])
                for cell in mtg_cells:
                    ws[cell].fill = gray_fill


if __name__ == "__main__":
    import argparse, os

    parser = argparse.ArgumentParser(
        description="Run Wochenplan pipeline: OGs (LA→non-LA/AA) → FR → Meetings."
    )
    parser.add_argument("-i", "--input",  required=True,
                        help=".xlsm source file (e.g. KW 41_leer.xlsm)")
    parser.add_argument("-o", "--output", required=True,
                        help="FINAL .xlsm (after meetings), e.g. KW_41_FINAL.xlsm")
    parser.add_argument("--csv", dest="csv_file", default=None,
                        help="CSV file with dienste to import (optional)")
    parser.add_argument("--seed", type=int, default=1234,
                        help="random seed for fair picks (default: 1234)")
    parser.add_argument("-nocleanup", "--no-cleanup", dest="no_cleanup",
                        action="store_true",
                        help="Skip pre-run cleanup of FR/OG/Meetings cells (default: cleanup runs)")
    parser.add_argument("--keep-intermediate", "--keep", dest="keep_intermediate",
                        action="store_true",
                        help="Keep intermediate .xlsm files (default: delete them)")

    args = parser.parse_args()

    # Start with fresh counters
    reset_all_counters()
    
    # Create single RNG for entire pipeline (reproducible with seed)
    rng = random.Random(args.seed)

    # Intermediate filenames (only used when --keep-intermediate is set)
    out_stem       = os.path.splitext(args.output)[0]
    cleaned_out    = f"{out_stem}_0_CLEANED.xlsm"
    og_leaders_out = f"{out_stem}_1_OG_LEADERS.xlsm"
    og_full_out    = f"{out_stem}_2_OG_FULL.xlsm"
    fr_out         = f"{out_stem}_3_FR_APPLIED.xlsm"

    # Load workbook once
    wb = load_workbook(args.input, data_only=False, keep_vba=True)
    ws = wb["Wochenplan"]

    # 0) Cleanup (default)
    if args.no_cleanup:
        cleaned_note = "SKIPPED"
    else:
        cleanup_blocks(ws, clear_fr=True, clear_og=True, clear_meetings=True)
        cleaned_note = cleaned_out
        if args.keep_intermediate:
            wb.save(cleaned_out)
            patch_xlsm(cleaned_out, args.input)

    # 0.5) Import dienste from CSV (if provided)
    if args.csv_file:
        print(f"📥 Importing dienste from CSV: {args.csv_file}")
        fill_dienste_from_csv(ws, args.csv_file)
        print("✓ CSV import complete")

    # Read absences once (cleanup does not touch absence cells)
    absences = read_absences_by_day(ws)

    # 1) OG leaders (LA)
    assign_la_to_ogs(ws, absences)
    if args.keep_intermediate:
        wb.save(og_leaders_out)
        patch_xlsm(og_leaders_out, args.input)

    # 2) OG non-leaders (FA/AA) + flags
    assign_nonleaders_to_ogs(ws, absences, rng)
    if args.keep_intermediate:
        wb.save(og_full_out)
        patch_xlsm(og_full_out, args.input)

    # 3) FR (depends on OG + Laufen in OG)
    assign_fr_shifts_to_cells(ws, absences, rng)
    if args.keep_intermediate:
        wb.save(fr_out)
        patch_xlsm(fr_out, args.input)

    # 4) Meetings
    assign_meetings(ws, absences, rng)
    
    # 5) Remove covers from visual absence list (purely cosmetic)
    remove_covers_from_absences_visual(ws)

    # Save final output once and patch
    wb.save(args.output)
    patch_xlsm(args.output, args.input)

    print("✔ Pipeline completed")
    print(f"  0_CLEANUP      → {cleaned_note}")
    if args.keep_intermediate:
        print(f"  1_OG_LA        → {og_leaders_out}")
        print(f"  2_OG_FULL      → {og_full_out}")
        print(f"  3_FR_APPLIED   → {fr_out}")
    print(f"  4_+MEETINGS    → {args.output}")

    # Print per-person statistics (FR + Rapporte)
    print_weekly_stats()

    if args.keep_intermediate:
        print("↪ Kept intermediates (user requested).")
