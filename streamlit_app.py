# -*- coding: utf-8 -*-
"""
Wochenplan Scheduler — Streamlit UI
Run: streamlit run streamlit_app.py
"""

import json
import os
import tempfile
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

import wochenplan_scheduler as sched

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

STAFF_JSON  = Path(__file__).parent / "staff.json"
LAYOUT_JSON = Path(__file__).parent / "layout.json"
MEETING_POOLS_JSON = Path(__file__).parent / "meeting_pools.json"
TEMPLATE_XLSM = Path(__file__).parent / "KW_xx_TEMPLATE.xlsm"
ALL_WEEKDAYS = sched.WEEKDAYS  # ["Montag","Dienstag","Mittwoch","Donnerstag","Freitag"]

# Pool constants
_POOL_TYPES_MAP = {
    "names": "Person",
    "group": "Gruppe",
    "spaetdienst_aa": "Spätdienst_AA",
    "hintergrund_vortag": "Hintergrund Vortag",
}
_POOL_TYPES = list(_POOL_TYPES_MAP.keys())
_POOL_TYPES_DISPLAY = list(_POOL_TYPES_MAP.values())

_GROUP_MAP = {
    "AA": "AA",
    "OA": "OA",
    "LA": "LA",
    "FA_ALL": "alle Fachärzte"
}
_GROUP_OPTIONS = list(_GROUP_MAP.keys())
_GROUP_DISPLAY = list(_GROUP_MAP.values())

_SITE_OPTIONS = ["BH", "LI"]

# ---------------------------------------------------------------------------
# Layout helpers
# ---------------------------------------------------------------------------

def load_layout() -> dict:
    with open(LAYOUT_JSON, encoding="utf-8") as f:
        return json.load(f)


def save_layout(data: dict) -> None:
    with open(LAYOUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    sched.load_layout_from_json(str(LAYOUT_JSON))


def _cells_to_str(cells) -> str:
    if isinstance(cells, (list, tuple)):
        return ", ".join(cells)
    return str(cells) if cells else ""


def _str_to_cells(s) -> list:
    if not isinstance(s, str) or not s.strip():
        return []
    return [c.strip() for c in s.replace(";", ",").split(",") if c.strip()]


# ---------------------------------------------------------------------------
# Staff persistence helpers
# ---------------------------------------------------------------------------

def save_staff_to_json() -> None:
    """Serialise current sched.staff_by_name to staff.json with site_rules structure."""
    records = [
        {
            "name": s.name,
            "role": s.role,
            "site": s.site,
            "leads_ogs": sorted(s.leads_ogs),
            "rotations": sorted(s.rotations),
            "avoid_ogs": sorted(s.avoid_ogs),
            "fr_excluded": s.fr_excluded,
            "fr_excluded_days": sorted(s.fr_excluded_days),
            "is_cover": s.is_cover,
        }
        for s in sched.staff_by_name.values()
    ]
    
    # Preserve site_rules from loaded data or use defaults
    site_rules = sched.SITE_RULES if sched.SITE_RULES else {
        "BH": {"no_oa_vormittag": False},
        "LI": {"no_oa_vormittag": True}
    }
    
    data = {
        "site_rules": site_rules,
        "staff": records
    }
    
    with open(STAFF_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def staff_to_display_dataframe() -> pd.DataFrame:
    """Read-only overview table shown above the edit form."""
    rows = []
    # Sort by surname (last part of name after last space)
    def get_surname(s):
        return s.name.split()[-1]
    
    for s in sorted(sched.staff_by_name.values(), key=get_surname):
        fr_info = "Immer" if s.fr_excluded else (
            ", ".join(d[:2] for d in sched.WEEKDAYS if d in s.fr_excluded_days) or "—"
        )
        rows.append({
            "Name": s.name,
            "Rolle": s.role,
            "Standort": s.site,
            "Organgruppenleitung": ", ".join(sorted(s.leads_ogs)) or "—",
            "Rotationen": ", ".join(sorted(s.rotations)) or "—",
            "Vermeiden": ", ".join(sorted(s.avoid_ogs)) or "—",
            "Kein Frontarzt": fr_info,
            "Stellvertreter": "Ja" if s.is_cover else "—",
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Rapporte helpers
# ---------------------------------------------------------------------------

_COL_CFG = {
    day: st.column_config.TextColumn(day[:2], help=day)
    for day in ALL_WEEKDAYS
}

def _rapport_overview_df() -> pd.DataFrame:
    """Overview table of all rapporte from meeting_pools.json."""
    pools_data = load_meeting_pools()
    rows = []
    for key in pools_data:
        if key.startswith("_"):
            continue  # settings key, not a rapport
        parts = key.split("|", 1)
        site = parts[0] if len(parts) == 2 else "?"
        name = parts[1] if len(parts) == 2 else key
        rows.append({"Key": key, "Name": name, "Standort": site})
    return pd.DataFrame(rows)


def _rename_rapport(old_key: str, new_key: str) -> None:
    """Rename a rapport key atomically in both layout.json and meeting_pools.json."""
    # --- meeting_pools.json ---
    pools_data = load_meeting_pools()
    if old_key not in pools_data:
        raise KeyError(f"Rapport '{old_key}' nicht in meeting_pools.json gefunden.")
    cfg = pools_data.pop(old_key)
    pools_data[new_key] = cfg
    save_meeting_pools(pools_data)

    # --- layout.json ---
    layout = load_layout()
    old_parts = old_key.split("|", 1)
    new_parts = new_key.split("|", 1)
    old_site, old_name = old_parts[0], old_parts[1]
    new_site, new_name = new_parts[0], new_parts[1]

    # Move cell mappings: remove from old site/name, insert under new site/name
    old_cells = layout["meeting_cells"].get(old_site, {}).pop(old_name, {})
    layout["meeting_cells"].setdefault(new_site, {})[new_name] = old_cells
    save_layout(layout)


def _delete_rapport(key: str) -> None:
    """Delete a rapport from both layout.json and meeting_pools.json."""
    pools_data = load_meeting_pools()
    pools_data.pop(key, None)
    save_meeting_pools(pools_data)

    layout = load_layout()
    parts = key.split("|", 1)
    if len(parts) == 2:
        layout["meeting_cells"].get(parts[0], {}).pop(parts[1], None)
    save_layout(layout)


def _add_rapport(key: str) -> None:
    """Add a new rapport to both layout.json and meeting_pools.json with empty defaults."""
    pools_data = load_meeting_pools()
    pools_data[key] = {
        "site": key.split("|", 1)[0],
        "pools": [{"type": "names", "names": [], "site": key.split("|", 1)[0]}],
        "fallback_text": "FÄLLT AUS",
        "roter_fallback_text": True,
    }
    save_meeting_pools(pools_data)

    layout = load_layout()
    parts = key.split("|", 1)
    site, name = parts[0], parts[1]
    layout["meeting_cells"].setdefault(site, {}).setdefault(name, {
        day: [] for day in ALL_WEEKDAYS
    })
    save_layout(layout)


# ---------------------------------------------------------------------------
# Pool helpers
# ---------------------------------------------------------------------------

def _exclude_if_day_to_str(eid: dict | None) -> str:
    """Convert {"Donnerstag": ["H.W. Ott"]} → 'Donnerstag: H.W. Ott'"""
    if not eid:
        return ""
    parts = []
    for day, names in eid.items():
        if isinstance(names, (list, tuple)):
            parts.append(f"{day}: {', '.join(names)}")
        else:
            parts.append(f"{day}: {names}")
    return "; ".join(parts)


def _str_to_exclude_if_day(s: str) -> dict | None:
    """Convert 'Donnerstag: H.W. Ott, X; Freitag: Y' → dict."""
    if not s or not s.strip():
        return None
    result = {}
    for part in s.split(";"):
        part = part.strip()
        if ":" not in part:
            continue
        day, names_str = part.split(":", 1)
        day = day.strip()
        names = [n.strip() for n in names_str.split(",") if n.strip()]
        if day and names:
            result[day] = names
    return result or None


def _list_to_str(lst: list | None) -> str:
    if not lst:
        return ""
    return ", ".join(str(x) for x in lst)


def _str_to_list(s: str) -> list:
    if not s or not s.strip():
        return []
    return [x.strip() for x in s.split(",") if x.strip()]


# ---------------------------------------------------------------------------
# Session-state bootstrap (runs once per browser session)
# ---------------------------------------------------------------------------

def _init_session_state() -> None:
    if "staff_loaded" not in st.session_state:
        # sched already loaded staff.json at import time if it existed;
        # nothing more to do — just mark the session as initialised.
        st.session_state["staff_loaded"] = True
    if "result_bytes" not in st.session_state:
        st.session_state["result_bytes"] = None
    if "result_filename" not in st.session_state:
        st.session_state["result_filename"] = "Wochenplan_FINAL.xlsm"


_init_session_state()

st.set_page_config(
    page_title="Wochenplan Scheduler",
    page_icon="📋",
    layout="wide",
)

# ---------------------------------------------------------------------------
# Password gate
# ---------------------------------------------------------------------------

def _check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True

    st.markdown(
        """
        <style>
        .login-box {
            max-width: 340px;
            margin: 8rem auto 0 auto;
            padding: 2rem 2rem 1.5rem 2rem;
            border-radius: 8px;
            background: #1a1a1a;
            box-shadow: 0 4px 24px rgba(0,0,0,0.5);
            text-align: center;
        }
        .login-title {
            font-size: 1.2rem;
            font-weight: 600;
            color: #e0e0e0;
            margin-bottom: 1.5rem;
            letter-spacing: 0.05em;
        }
        </style>
        <div class="login-box">
            <div class="login-title">🔒 Wochenplan Scheduler</div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            pw = st.text_input("Passwort", type="password", label_visibility="collapsed",
                               placeholder="Passwort eingeben")
            submitted = st.form_submit_button("Anmelden", use_container_width=True, type="primary")
            
        if submitted:
            if pw == st.secrets.get("password", ""):
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Falsches Passwort.")
    return False

if not _check_password():
    st.stop()

st.title("Wochenplan Scheduler — KSBL Radiologie")
st.caption("by S. Vitéz · Powered by Anthropic")

# Helper functions for meeting pools (used in multiple tabs)
def load_meeting_pools() -> dict:
    with open(MEETING_POOLS_JSON, encoding="utf-8") as f:
        return json.load(f)

def save_meeting_pools(data: dict) -> None:
    with open(MEETING_POOLS_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    sched.load_meeting_pools_from_json(str(MEETING_POOLS_JSON))

# Helper functions for og_rules (used in Organgruppen tab)
def load_og_rules() -> dict:
    og_rules_path = Path(__file__).parent / "og_rules.json"
    with open(og_rules_path, encoding="utf-8") as f:
        return json.load(f)

def save_og_rules(data: dict) -> None:
    og_rules_path = Path(__file__).parent / "og_rules.json"
    with open(og_rules_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    # Reload all og_rules derived variables in scheduler module
    sched.OG_ROTATION_OR_LEADER_ONLY, sched.OG_WARN_KEIN_AA, sched.TARGET_OG_FOR_ONE_FA, sched.TARGET_OG_FOR_KEIN_FA_SITE, sched.OG_EXCLUDE_FROM_RAPPORTE = sched._load_og_rules()
    sched.OGS_SKIP_KEIN_AA = set(sched.OG_LIST) - sched.OG_WARN_KEIN_AA


# ---------------------------------------------------------------------------
# Known Bezeichnungen — loaded from bezeichnungen.json (single source of truth)
# Edit bezeichnungen.json to add new types; no changes needed here.
# ---------------------------------------------------------------------------
def _load_known_bezeichnungen() -> set:
    _path = Path(__file__).parent / "bezeichnungen.json"
    with open(_path, encoding="utf-8") as _f:
        _bez = json.load(_f)
    known = set()
    for key, values in _bez.items():
        if not key.startswith("_"):
            known.update(values)
    return known

_KNOWN_BEZEICHNUNGEN = _load_known_bezeichnungen()


def _check_unknown_bezeichnungen(csv_path: str) -> list[str]:
    """Scan CSV for Bezeichnungen not recognized by the catalog. A Bezeichnung is
    recognized if any catalog token appears as a substring of it — identical to
    the scheduler's substring matching. Returns list of unrecognized strings."""
    import csv as _csv
    unknown = []
    for encoding in ("utf-8", "iso-8859-1"):
        try:
            with open(csv_path, encoding=encoding, newline="") as f:
                reader = _csv.DictReader(f, delimiter=";")
                for row in reader:
                    b = row.get("Bezeichnung", "").strip()
                    if not b or b in unknown:
                        continue
                    recognized = any(tok in b for tok in _KNOWN_BEZEICHNUNGEN)
                    if not recognized:
                        unknown.append(b)
            break
        except (UnicodeDecodeError, KeyError):
            continue
    return unknown

def _staff_form(form_key: str, defaults: dict | None = None) -> dict | None:
    """Render a staff edit/add form and return the submitted values as a dict,
    or None if the form was not submitted."""
    d = defaults or {}
    with st.form(form_key, clear_on_submit=(defaults is None)):
        c1, c2, c3 = st.columns(3)
        name     = c1.text_input("Name", value=d.get("name", ""), placeholder="J. Beispiel",
                                 disabled=(defaults is not None))
        role     = c2.selectbox("Rolle", ["AA", "OA", "LA"],
                                index=["AA", "OA", "LA"].index(d.get("role", "AA")))
        site     = c3.selectbox("Standort", ["BH", "LI"],
                                index=["BH", "LI"].index(d.get("site", "BH")))

        c4, c5 = st.columns(2)
        leads = c4.multiselect(
            "Organgruppenleitung",
            options=sched.OG_LIST,
            default=sorted(d.get("leads_ogs", [])),
            help="Nur Auswählen falls zutreffend. Laufen wird wie eine Organgruppenleitung gehandhabt",
        )
        rots = c5.multiselect(
            "Rotationen",
            options=sched.OG_LIST,
            default=sorted(d.get("rotations", [])),
        )

        avoid = st.multiselect(
            "Organgruppe vermeiden",
            options=sched.OG_LIST,
            default=sorted(d.get("avoid_ogs", [])),
            help="Diese Person wird diesen Organgruppen nur als letzte Möglichkeit zugeteilt "
                 "(wenn niemand sonst verfügbar ist). Nur für OA wirksam.",
        )

        st.markdown("**Kein Frontarzt**")
        fr_col1, fr_col2 = st.columns([1, 2])
        # Only disable for AA when editing (defaults is not None)
        disable_for_aa = (role == "AA" and defaults is not None)
        fr_always = fr_col1.checkbox(
            "Nie Frontarzt",
            value=d.get("fr_excluded", False),
            disabled=disable_for_aa,
            help="Nur relevant für Fachärzte (OA/LA)." if disable_for_aa else None,
        )
        fr_days = fr_col2.multiselect(
            "Nur an diesen Tagen kein Frontarzt",
            options=sched.WEEKDAYS,
            default=sorted(d.get("fr_excluded_days", [])),
            disabled=(fr_always or disable_for_aa),
            help="Wird ignoriert wenn 'Nie Frontarzt' aktiviert ist oder Person kein Facharzt ist.",
        )

        st.markdown("**Stellvertretungsregel**")
        is_cover = st.checkbox(
            "Ist Stellvertreter",
            value=d.get("is_cover", False),
            help="Stellvertreter werden nicht in der wöchentlichen Abwesenheitsliste aufgeführt, auch wenn sie abwesend sind",
        )

        label = "Änderungen speichern" if defaults else "Hinzufügen"
        submitted = st.form_submit_button(label, type="primary")
        if submitted:
            return {
                "name": name.strip(),
                "role": role,
                "site": site,
                "leads_ogs": leads if role == "LA" else [],
                "rotations": rots,
                "avoid_ogs": avoid,
                "fr_excluded": fr_always,
                "fr_excluded_days": [] if fr_always else fr_days,
                "is_cover": is_cover,
            }
    return None


# ===========================================================================
# SIDEBAR NAVIGATION
# ===========================================================================

st.sidebar.title("Navigation")

page = st.sidebar.radio(
    "Seite auswählen:",
    [
        "📋 Wochenplan (Standard)",
        "📄 Wochenplan (Eigene Vorlage)", 
        "👥 Personalverwaltung",
        "📊 Rapporte verwalten",
        "📊 Rapporte-Pools",
        "📈 Rapport-Statistik",
        "👨‍⚕️ Frontarzt",
        "🏥 Organgruppen Verwalten",
        "🏥 Organgruppen Regeln",
        "🔧 Layout-Editor"
    ],
    label_visibility="collapsed"
)

# ===========================================================================
# PAGE 1 — Wochenplan (Standard)
# ===========================================================================

if page == "📋 Wochenplan (Standard)":
    # Check if template exists
    template_exists = TEMPLATE_XLSM.exists()
    if not template_exists:
        st.warning("⚠️ Keine 'KW_xx_TEMPLATE.xlsm' gefunden. Bitte lade eine Vorlage hoch (unten).")
    
    st.markdown("### 📤 CSV + Standard-Vorlage 🪄→ fertiger Wochenplan")
    
    csv_file_opt1 = st.file_uploader(
        "CSV-Datei hochladen",
        type=["csv"],
        help="Lade nur eine CSV-Datei hoch. Die Standard-Vorlage wird automatisch geladen.",
        key="csv_opt1",
        disabled=not template_exists
    )
    st.caption("*Für Wochenenddienste: CSV bitte ab Samstag der Vorwoche beginnen lassen, damit Hintergrunddienste korrekt berücksichtigt werden.*")
    
    col1, col2 = st.columns([1, 2])
    with col1:
        seed_opt1 = st.number_input(
            "Seed", value=1234, step=1, format="%d",
            help="Zufalls-Seed für reproduzierbare Ergebnisse.",
            key="seed_opt1"
        )
    
    skip_stats_opt1 = st.checkbox(
        "Keine Statistik speichern (Testlauf)",
        value=False,
        key="skip_stats_opt1",
    )
    run_opt1 = st.button(
        "Wochenplan erstellen (mit Standard-Vorlage)",
        disabled=(csv_file_opt1 is None or not template_exists),
        type="primary",
        key="run_opt1"
    )
    
    if run_opt1 and csv_file_opt1 and template_exists:
        csv_tmp_path = None
        output_tmp_path = None
        try:
            # Save CSV to temp
            with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as csv_tmp:
                csv_tmp.write(csv_file_opt1.getbuffer())
                csv_tmp_path = csv_tmp.name
            
            with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as f_out:
                output_tmp_path = f_out.name
            
            with st.spinner("Pipeline läuft…"):
                
                # Reset counters
                sched.reset_all_counters()
                
                # Load template from disk
                wb = load_workbook(str(TEMPLATE_XLSM), data_only=False, keep_vba=True)
                ws = wb["Wochenplan"]
                
                # Stage 0: Cleanup
                sched.cleanup_blocks(ws, clear_fr=True, clear_og=True, clear_meetings=True)
                
                # Stage 0.5: CSV import
                unknown_bez = _check_unknown_bezeichnungen(csv_tmp_path)
                if unknown_bez:
                    st.warning(
                        "⚠️ **Unbekannte Bezeichnung(en) im CSV** — betroffene Personen werden als anwesend behandelt:\n\n"
                        + "\n".join(f"- `{b}`" for b in unknown_bez)
                        + "\n\nBitte `bezeichnungen.json` ergänzen falls nötig."
                    )
                sched.fill_dienste_from_csv(ws, csv_tmp_path)
                
                # Read absences
                absences = sched.read_absences_by_day(ws)
                
                # Create RNG from seed for reproducibility
                import random
                rng = random.Random(seed_opt1)
                
                # Stage 1: OG
                sched.assign_la_to_ogs(ws, absences)
                sched.assign_nonleaders_to_ogs(ws, absences, rng)
                
                # Stage 2: FR
                sched.assign_fr_shifts_to_cells(ws, absences, rng)
                
                # Stage 3: Meetings
                sched.assign_meetings(ws, absences, rng, skip_stats=skip_stats_opt1)

                # Write stats history to Statistik sheet
                sched.write_stats_to_sheet(wb)

                # Save
                wb.save(output_tmp_path)
                wb.close()
            
            # Offer download
            with open(output_tmp_path, "rb") as f:
                st.download_button(
                    "⬇️ Finaler Wochenplan herunterladen",
                    data=f.read(),
                    file_name="Wochenplan_FINAL.xlsm",
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                )
            
            st.success("✔ Pipeline erfolgreich abgeschlossen!")
            sched.print_weekly_stats()
            
        except Exception as e:
            st.error(f"Fehler: {e}")
            import traceback
            st.code(traceback.format_exc())
        finally:
            if csv_tmp_path and os.path.exists(csv_tmp_path):
                os.unlink(csv_tmp_path)
            if output_tmp_path and os.path.exists(output_tmp_path):
                os.unlink(output_tmp_path)
    

    st.divider()
    
    st.markdown("### 🗂️ Vorlage verwalten")
    st.caption("Lade eine neue leere Vorlage hoch, um die Standard-Vorlage zu ersetzen.")
    
    with st.expander("Neue Vorlage hochladen"):
        
        new_template = st.file_uploader(
            "Neue Vorlage hochladen",
            type=["xlsm"],
            help="Die neue leere Wochenplan-Vorlage. Wird automatisch als 'KW_xx_TEMPLATE.xlsm' gespeichert.",
            key="new_template_uploader"
        )
        
        if new_template:
            if st.button("Vorlage ersetzen", type="primary", key="replace_template_btn"):
                try:
                    with open(TEMPLATE_XLSM, "wb") as f:
                        f.write(new_template.getbuffer())
                    st.success(f"✓ '{new_template.name}' wurde als 'KW_xx_TEMPLATE.xlsm' gespeichert!")
                except Exception as e:
                    st.error(f"Fehler beim Ersetzen: {e}")

    # Feiertage section
    st.divider()
    st.markdown("### 🎉 Feiertage")
    st.caption("Wähle Wochentage, die als Feiertage behandelt werden sollen (keine Absenzen, OG, FR, Rapporte).")
    
    # Load current feiertage from layout
    layout_feiertage = load_layout()
    current_feiertage = layout_feiertage.get("feiertage", [])
    
    feiertage_selected = st.multiselect(
        "Feiertage",
        options=["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"],
        default=current_feiertage,
        help="An diesen Tagen werden nur Nacht-, Hintergrund- und Vordergrunddienste eingeplant.",
        key="feiertage_select_standard"
    )
    
    if st.button("Feiertage speichern", type="primary", key="save_feiertage_standard"):
        layout_feiertage["feiertage"] = feiertage_selected
        save_layout(layout_feiertage)
        st.success(f"✓ Feiertage gespeichert: {', '.join(feiertage_selected) if feiertage_selected else 'Keine'}")
        st.rerun()

    
    
# ===========================================================================
# TAB 2 — Wochenplan mit eigener Vorlage
# ===========================================================================
 
elif page == "📄 Wochenplan (Eigene Vorlage)":

    st.markdown("### 📤 CSV + 📤 Eigene Vorlage 🪄→ fertiger Wochenplan")
    
    csv_file_opt2 = st.file_uploader(
        "CSV-Datei hochladen",
        type=["csv"],
        help="CSV-Datei mit Absenzen und Diensten.",
        key="csv_opt2"
    )
    st.caption("*Für Wochenenddienste: CSV bitte ab Samstag der Vorwoche beginnen lassen, damit Hintergrunddienste korrekt berücksichtigt werden.*")
    
    template_file_opt2 = st.file_uploader(
        "Eigene Wochenplan-Vorlage (.xlsm) hochladen",
        type=["xlsm"],
        help="Lade eine eigene .xlsm-Vorlage hoch.",
        key="template_opt2"
    )
    
    col1, col2 = st.columns([1, 2])
    with col1:
        seed_opt2 = st.number_input(
            "Seed", value=1234, step=1, format="%d",
            help="Zufalls-Seed für reproduzierbare Ergebnisse.",
            key="seed_opt2"
        )
    
    skip_stats_opt2 = st.checkbox(
        "Keine Statistik speichern (Testlauf)",
        value=False,
        key="skip_stats_opt2",
    )
    run_opt2 = st.button(
        "Wochenplan erstellen (mit eigener Vorlage)",
        disabled=(csv_file_opt2 is None or template_file_opt2 is None),
        type="primary",
        key="run_opt2"
    )
    
    if run_opt2 and csv_file_opt2 and template_file_opt2:
        csv_tmp_path = None
        template_tmp_path = None
        output_tmp_path = None
        try:
            # Save files to temp
            with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as csv_tmp:
                csv_tmp.write(csv_file_opt2.getbuffer())
                csv_tmp_path = csv_tmp.name
            
            with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as tmpl_tmp:
                tmpl_tmp.write(template_file_opt2.getbuffer())
                template_tmp_path = tmpl_tmp.name
            
            with tempfile.NamedTemporaryFile(suffix=".xlsm", delete=False) as f_out:
                output_tmp_path = f_out.name
            
            with st.spinner("Pipeline läuft…"):
                # Configure Laufen days
                # Reset counters
                sched.reset_all_counters()
                
                # Load uploaded template
                wb = load_workbook(template_tmp_path, data_only=False, keep_vba=True)
                ws = wb["Wochenplan"]
                
                # Stage 0: Cleanup
                sched.cleanup_blocks(ws, clear_fr=True, clear_og=True, clear_meetings=True)
                
                # Stage 0.5: CSV import
                unknown_bez = _check_unknown_bezeichnungen(csv_tmp_path)
                if unknown_bez:
                    st.warning(
                        "⚠️ **Unbekannte Bezeichnung(en) im CSV** — betroffene Personen werden als anwesend behandelt:\n\n"
                        + "\n".join(f"- `{b}`" for b in unknown_bez)
                        + "\n\nBitte `bezeichnungen.json` ergänzen falls nötig."
                    )
                sched.fill_dienste_from_csv(ws, csv_tmp_path)
                
                # Read absences
                absences = sched.read_absences_by_day(ws)
                
                # Create RNG from seed for reproducibility
                import random
                rng = random.Random(seed_opt2)
                
                # Stage 1: OG
                sched.assign_la_to_ogs(ws, absences)
                sched.assign_nonleaders_to_ogs(ws, absences, rng)
                
                # Stage 2: FR
                sched.assign_fr_shifts_to_cells(ws, absences, rng)
                
                # Stage 3: Meetings
                sched.assign_meetings(ws, absences, rng, skip_stats=skip_stats_opt2)

                # Write stats history to Statistik sheet
                sched.write_stats_to_sheet(wb)

                # Save
                wb.save(output_tmp_path)
                wb.close()
            
            # Offer download
            with open(output_tmp_path, "rb") as f:
                st.download_button(
                    "⬇️ Finaler Wochenplan herunterladen",
                    data=f.read(),
                    file_name="Wochenplan_FINAL.xlsm",
                    mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                )
            
            st.success("✔ Pipeline erfolgreich abgeschlossen!")
            sched.print_weekly_stats()
            
        except Exception as e:
            st.error(f"Fehler: {e}")
            import traceback
            st.code(traceback.format_exc())
        finally:
            if csv_tmp_path and os.path.exists(csv_tmp_path):
                os.unlink(csv_tmp_path)
            if template_tmp_path and os.path.exists(template_tmp_path):
                os.unlink(template_tmp_path)
            if output_tmp_path and os.path.exists(output_tmp_path):
                os.unlink(output_tmp_path)
    
    # Feiertage section
    st.divider()
    st.markdown("### 🎉 Feiertage")
    st.caption("Wähle Wochentage, die als Feiertage behandelt werden sollen (keine Absenzen, OG, FR, Rapporte).")
    
    # Load current feiertage from layout
    layout_feiertage_eigene = load_layout()
    current_feiertage_eigene = layout_feiertage_eigene.get("feiertage", [])
    
    feiertage_selected_eigene = st.multiselect(
        "Feiertage",
        options=["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"],
        default=current_feiertage_eigene,
        help="An diesen Tagen werden nur Nacht-, Hintergrund- und Vordergrunddienste eingeplant.",
        key="feiertage_select_eigene"
    )
    
    if st.button("Feiertage speichern", type="primary", key="save_feiertage_eigene"):
        layout_feiertage_eigene["feiertage"] = feiertage_selected_eigene
        save_layout(layout_feiertage_eigene)
        st.success(f"✓ Feiertage gespeichert: {', '.join(feiertage_selected_eigene) if feiertage_selected_eigene else 'Keine'}")
        st.rerun()

# ===========================================================================
# TAB 4 — Personalverwaltung
# ===========================================================================
elif page == "👥 Personalverwaltung":
    st.subheader("Personalbestand")

    # Read-only overview table with color-coded roles
    df = staff_to_display_dataframe()
    
    # Define role colors (softer, 30% opacity)
    def color_row(row):
        colors = {
            "AA": "background-color: rgba(144, 238, 144, 0.3)",  # light green
            "OA": "background-color: rgba(135, 206, 235, 0.3)",  # sky blue
            "LA": "background-color: rgba(255, 182, 198, 0.3)",  # light pink
        }
        color = colors.get(row["Rolle"], "")
        return [color] * len(row)
    
    # Apply styling to entire rows
    styled_df = df.style.apply(color_row, axis=1)
    
    st.dataframe(
        styled_df,
        use_container_width=True,
        hide_index=True,
    )

    st.divider()

    # ---- Edit existing staff member ----
    st.subheader("Mitarbeiter bearbeiten")
    all_names = sorted(sched.staff_by_name.keys())
    edit_name = st.selectbox("Person auswählen", all_names, key="edit_select")

    if edit_name:
        s = sched.staff_by_name[edit_name]
        edit_defaults = {
            "name": s.name,
            "role": s.role,
            "site": s.site,
            "leads_ogs": sorted(s.leads_ogs),
            "rotations": sorted(s.rotations),
            "avoid_ogs": sorted(s.avoid_ogs),
            "fr_excluded": s.fr_excluded,
            "fr_excluded_days": sorted(s.fr_excluded_days),
            "is_cover": s.is_cover,
        }
        result = _staff_form("edit_staff_form", defaults=edit_defaults)
        if result:
            sched.add_staff(
                name=edit_name,
                role=result["role"],
                site=result["site"],
                leads_for=result["leads_ogs"],
                rotation=result["rotations"],
                avoid=result["avoid_ogs"],
                fr_excluded=result["fr_excluded"],
                fr_excluded_days=result["fr_excluded_days"],
                is_cover=result["is_cover"],
            )
            sched.rebuild_quick_views()
            save_staff_to_json()
            st.success(f"'{edit_name}' aktualisiert und gespeichert.")
            st.rerun()

        if st.button("Mitarbeiter löschen", key="delete_btn"):
            deleted_name = edit_name
            del sched.staff_by_name[deleted_name]
            sched.rebuild_quick_views()
            save_staff_to_json()
            
            # Clean up orphaned names from meeting_pools.json
            pools_data = load_meeting_pools()
            pools_modified = False
            all_current_names = set(sched.staff_by_name.keys())
            
            for meeting_key, cfg in pools_data.items():
                if meeting_key.startswith("_"):
                    continue
                for pool in cfg.get("pools", []):
                    # Clean names list
                    if "names" in pool and pool["names"]:
                        original_names = pool["names"]
                        pool["names"] = [n for n in original_names if n in all_current_names]
                        if len(pool["names"]) != len(original_names):
                            pools_modified = True
                    
                    # Clean exclude_names
                    if "exclude_names" in pool and pool["exclude_names"]:
                        original_excluded = pool["exclude_names"]
                        pool["exclude_names"] = [n for n in original_excluded if n in all_current_names]
                        if len(pool["exclude_names"]) != len(original_excluded):
                            pools_modified = True
                        if not pool["exclude_names"]:
                            pool["exclude_names"] = None
                    
                    # Clean exclude_if_day
                    if "exclude_if_day" in pool and pool["exclude_if_day"]:
                        for day, names in list(pool["exclude_if_day"].items()):
                            cleaned = [n for n in names if n in all_current_names]
                            if cleaned:
                                pool["exclude_if_day"][day] = cleaned
                            else:
                                del pool["exclude_if_day"][day]
                            if cleaned != names:
                                pools_modified = True
                        if not pool["exclude_if_day"]:
                            pool["exclude_if_day"] = None
            
            if pools_modified:
                save_meeting_pools(pools_data)
            
            st.success(f"'{deleted_name}' wurde entfernt" + (" (und aus Rapporte-Pools entfernt)." if pools_modified else "."))
            st.rerun()

    st.divider()

    # ---- Add new staff member ----
    with st.expander("Neues Personalmitglied hinzufügen"):
        result = _staff_form("add_staff_form")
        if result:
            name_clean = result["name"]
            if not name_clean:
                st.warning("Name darf nicht leer sein.")
            elif name_clean in sched.staff_by_name:
                st.warning(f"'{name_clean}' ist bereits im Personalbestand.")
            else:
                sched.add_staff(
                    name=name_clean,
                    role=result["role"],
                    site=result["site"],
                    leads_for=result["leads_ogs"],
                    rotation=result["rotations"],
                    avoid=result["avoid_ogs"],
                    fr_excluded=result["fr_excluded"],
                    fr_excluded_days=result["fr_excluded_days"],
                    is_cover=result["is_cover"],
                )
                sched.rebuild_quick_views()
                save_staff_to_json()
                st.success(f"'{name_clean}' wurde hinzugefügt und gespeichert.")
                st.rerun()


elif page == "📊 Rapporte verwalten":
    st.subheader("Rapporte verwalten")
    st.caption("Rapporte hinzufügen, umbenennen, löschen oder neu anordnen. Zellen werden im Layout-Editor bearbeitet. "
               "**Die Reihenfolge bestimmt die Zuteilungsreihenfolge:** Rapporte werden von oben nach unten zugeteilt. "
               "Rapporte mit Statistik sollten in der Regel vor den übrigen stehen, damit ihre Zuteilung in die "
               "Wochenauslastung der späteren Rapporte einfließt.")

    pools_data = load_meeting_pools()

    # ---- Global statistics setting ----
    _settings = pools_data.get("_settings", {}) if isinstance(pools_data.get("_settings"), dict) else {}
    respect_inweek = st.checkbox(
        "Statistik-Rapporte berücksichtigen die Wochenauslastung",
        value=bool(_settings.get("stats_respect_inweek", True)),
        key="stats_respect_inweek_cb",
        help="Wenn aktiviert, werden Personen, die in dieser Woche bereits andere Rapporte haben, bei "
             "Statistik-Rapporten zurückgestellt (zuerst Tages-, dann Wochenauslastung, dann wochenübergreifende "
             "Statistik). Wenn deaktiviert, zählt nur die wochenübergreifende Statistik.",
    )
    if bool(_settings.get("stats_respect_inweek", True)) != respect_inweek:
        _settings["stats_respect_inweek"] = respect_inweek
        pools_data["_settings"] = _settings
        save_meeting_pools(pools_data)
        st.rerun()

    st.divider()

    # ---- Reorderable overview ----
    st.markdown("**Reihenfolge der Rapporte** (↑/↓ zum Verschieben)")
    ordered_keys = [k for k in pools_data if not k.startswith("_")]
    for pos, key in enumerate(ordered_keys):
        parts = key.split("|", 1)
        site = parts[0] if len(parts) == 2 else "?"
        name = parts[1] if len(parts) == 2 else key
        stat_mark = " 📊" if pools_data[key].get("statistik_führen") else ""
        num_col, name_col, up_col, down_col = st.columns([1, 8, 1, 1])
        num_col.markdown(f"**{pos+1}**")
        name_col.markdown(f"`{site}` {name}{stat_mark}")
        if pos > 0:
            if up_col.button("↑", key=f"rap_up_{pos}"):
                kk = list(pools_data.keys())
                i = kk.index(key)
                kk[i], kk[i-1] = kk[i-1], kk[i]
                pools_data = {k: pools_data[k] for k in kk}
                save_meeting_pools(pools_data); st.rerun()
        if pos < len(ordered_keys) - 1:
            if down_col.button("↓", key=f"rap_down_{pos}"):
                kk = list(pools_data.keys())
                i = kk.index(key)
                kk[i], kk[i+1] = kk[i+1], kk[i]
                pools_data = {k: pools_data[k] for k in kk}
                save_meeting_pools(pools_data); st.rerun()

    st.divider()

    rapport_df = _rapport_overview_df()

    st.divider()

    # ---- Edit existing rapport ----
    st.subheader("Rapport bearbeiten")
    all_rapport_keys = rapport_df["Key"].tolist() if not rapport_df.empty else []
    edit_rapport_key = st.selectbox("Rapport auswählen", all_rapport_keys, key="edit_rapport_select")

    if edit_rapport_key:
        r_parts = edit_rapport_key.split("|", 1)
        r_site_current = r_parts[0]
        r_name_current = r_parts[1] if len(r_parts) == 2 else edit_rapport_key

        rc1, rc2 = st.columns(2)
        r_name_new = rc1.text_input("Name", value=r_name_current, key="edit_rapport_name")
        r_site_new = rc2.selectbox("Standort", ["BH", "LI"],
                                   index=["BH", "LI"].index(r_site_current) if r_site_current in ["BH", "LI"] else 0,
                                   key="edit_rapport_site")

        if st.button("Änderungen speichern", type="primary", key="save_rapport_btn"):
            new_key = f"{r_site_new}|{r_name_new.strip()}"
            if not r_name_new.strip():
                st.warning("Name darf nicht leer sein.")
            elif new_key != edit_rapport_key and new_key in all_rapport_keys:
                st.warning(f"'{new_key}' existiert bereits.")
            else:
                try:
                    _rename_rapport(edit_rapport_key, new_key)
                    st.success(f"Rapport umbenannt zu '{new_key}'.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Fehler: {e}")

        if st.button("Rapport löschen", key="delete_rapport_btn"):
            try:
                _delete_rapport(edit_rapport_key)
                st.success(f"'{edit_rapport_key}' wurde gelöscht.")
                st.rerun()
            except Exception as e:
                st.error(f"Fehler: {e}")

    st.divider()

    # ---- Add new rapport ----
    with st.expander("Neuen Rapport hinzufügen"):
        nc1, nc2 = st.columns(2)
        new_r_name = nc1.text_input("Name", placeholder="z.B. Medizin (07:45-08:00)", key="new_rapport_name")
        new_r_site = nc2.selectbox("Standort", ["BH", "LI"], key="new_rapport_site")

        if st.button("Rapport hinzufügen", type="primary", key="add_rapport_btn"):
            new_r_key = f"{new_r_site}|{new_r_name.strip()}"
            if not new_r_name.strip():
                st.warning("Name darf nicht leer sein.")
            elif new_r_key in all_rapport_keys:
                st.warning(f"'{new_r_key}' existiert bereits.")
            else:
                try:
                    _add_rapport(new_r_key)
                    st.success(f"'{new_r_key}' wurde hinzugefügt. Zellen im Layout-Editor eintragen.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Fehler: {e}")


elif page == "🔧 Layout-Editor":
    st.subheader("Layout-Editor")
    st.caption(
        "Excel-Zellreferenzen für alle Planabschnitte bearbeiten. "
        "Mehrere Zellen kommagetrennt eingeben, z.B. «T35, T36, T37»."
    )

    layout = load_layout()

    # ---- Abwesenheiten & Nachtdienst ----
    with st.expander("Abwesenheiten & Nachtdienst-Bereiche"):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Abwesenheiten** (Bereich je Tag)")
            abw_df = pd.DataFrame(
                [{"Tag": d, "Bereich": layout["abw_ranges"].get(d, "")} for d in ALL_WEEKDAYS]
            ).set_index("Tag")
            abw_edited = st.data_editor(abw_df, key="abw_ed", use_container_width=True)
        with c2:
            st.markdown("**Nachtdienst** (Bereich je Tag)")
            nacht_df = pd.DataFrame(
                [{"Tag": d, "Bereich": layout["nacht_ranges"].get(d, "")} for d in ALL_WEEKDAYS]
            ).set_index("Tag")
            nacht_edited = st.data_editor(nacht_df, key="nacht_ed", use_container_width=True)

    # ---- Spätdienst ----
    with st.expander("Spätdienst-Zellen"):
        spaet_rows = []
        for site in ["BH", "LI"]:
            row = {"Standort": site}
            for day in ALL_WEEKDAYS:
                row[day] = layout["spaetdienst_cells"].get(site, {}).get(day, "")
            spaet_rows.append(row)
        spaet_df = pd.DataFrame(spaet_rows).set_index("Standort")
        spaet_edited = st.data_editor(
            spaet_df, key="spaet_ed", use_container_width=True, column_config=_COL_CFG
        )

    # ---- Frontarzt ----
    with st.expander("Frontarzt-Zellen"):
        fr_rows = []
        for site in ["BH", "LI"]:
            row = {"Standort": site}
            for day in ALL_WEEKDAYS:
                row[day] = _cells_to_str(layout["fr_cells"].get(site, {}).get(day, []))
            fr_rows.append(row)
        fr_df = pd.DataFrame(fr_rows).set_index("Standort")
        fr_edited = st.data_editor(
            fr_df, key="fr_ed", use_container_width=True, column_config=_COL_CFG
        )

    # ---- Organgruppen ----
    with st.expander("Organgruppen-Zellen"):
        og_rows = []
        for og in sched.OG_LIST:
            row = {"OG": og}
            for day in ALL_WEEKDAYS:
                row[day] = _cells_to_str(layout["og_cells"].get(og, {}).get(day, []))
            og_rows.append(row)
        og_df = pd.DataFrame(og_rows).set_index("OG")
        og_edited = st.data_editor(
            og_df, key="og_ed", use_container_width=True, column_config=_COL_CFG
        )

    # ---- Rapporte BH ----
    with st.expander("Rapporte-Zellen — BH"):
        bh_rows = []
        for mtg, day_map in layout["meeting_cells"].get("BH", {}).items():
            row = {"Rapport": mtg}
            for day in ALL_WEEKDAYS:
                row[day] = _cells_to_str(day_map.get(day, []))
            bh_rows.append(row)
        bh_df = pd.DataFrame(bh_rows).set_index("Rapport")
        bh_edited = st.data_editor(
            bh_df, key="bh_mtg_ed", use_container_width=True, column_config=_COL_CFG
        )

    # ---- Rapporte LI ----
    with st.expander("Rapporte-Zellen — LI"):
        li_rows = []
        for mtg, day_map in layout["meeting_cells"].get("LI", {}).items():
            row = {"Rapport": mtg}
            for day in ALL_WEEKDAYS:
                row[day] = _cells_to_str(day_map.get(day, []))
            li_rows.append(row)
        li_df = pd.DataFrame(li_rows).set_index("Rapport")
        li_edited = st.data_editor(
            li_df, key="li_mtg_ed", use_container_width=True, column_config=_COL_CFG
        )

    # ---- Medizin Montag placeholder ----
    with st.expander("Medizin Montag Platzhalter"):
        st.caption("Zelle für «BITTE EINTRAGEN» am Montag je Standort")
        _mc = layout.get("medizin_monday_cells", {})
        mc1, mc2 = st.columns(2)
        medizin_bh = mc1.text_input("BH", value=_mc.get("BH", ""), key="medizin_bh")
        medizin_li = mc2.text_input("LI", value=_mc.get("LI", ""), key="medizin_li")

    # ---- Feiertage Merge Cells ----
    with st.expander("Zusammenführen von Zellen bei Feiertagen"):
        st.caption("Zellbereiche für Vordergrunddienst an Feiertagen (z.B. T23:AC24)")
        _fmc = layout.get("feiertage_merge_cells", {})
        
        fmc_rows = []
        for day in ALL_WEEKDAYS:
            fmc_rows.append({"Tag": day, "Bereich": _fmc.get(day, "")})
        
        fmc_df = pd.DataFrame(fmc_rows).set_index("Tag")
        feiertage_merge_edited = st.data_editor(
            fmc_df, 
            key="feiertage_merge_ed", 
            use_container_width=True,
            height=250
        )

    # ---- Save ----
    st.divider()
    if st.button("Alle Änderungen speichern", type="primary", key="save_layout_btn"):
        # Load current layout to preserve fields not in the editor
        current_layout = load_layout()
        
        new_layout = {
            "abw_ranges": {
                row: abw_edited.at[row, "Bereich"] for row in abw_edited.index
            },
            "nacht_ranges": {
                # Save edited weekdays (Mon-Fri)
                **{row: nacht_edited.at[row, "Bereich"] for row in nacht_edited.index},
                # Preserve weekend from current layout
                "Samstag": current_layout.get("nacht_ranges", {}).get("Samstag", ""),
                "Sonntag": current_layout.get("nacht_ranges", {}).get("Sonntag", ""),
            },
            "spaetdienst_cells": {
                site: {day: str(spaet_edited.at[site, day] or "") for day in ALL_WEEKDAYS}
                for site in ["BH", "LI"]
            },
            "fr_cells": {
                site: {day: _str_to_cells(fr_edited.at[site, day]) for day in ALL_WEEKDAYS}
                for site in ["BH", "LI"]
            },
            "og_cells": {
                og: {day: _str_to_cells(og_edited.at[og, day]) for day in ALL_WEEKDAYS}
                for og in sched.OG_LIST
            },
            "meeting_cells": {
                "BH": {
                    mtg: {day: _str_to_cells(bh_edited.at[mtg, day]) for day in ALL_WEEKDAYS}
                    for mtg in bh_edited.index
                },
                "LI": {
                    mtg: {day: _str_to_cells(li_edited.at[mtg, day]) for day in ALL_WEEKDAYS}
                    for mtg in li_edited.index
                },
            },
            "medizin_monday_cells": {
                "BH": medizin_bh.strip(),
                "LI": medizin_li.strip(),
            },
            "feiertage_merge_cells": {
                row: feiertage_merge_edited.at[row, "Bereich"] for row in feiertage_merge_edited.index
            },
            # Preserve new CSV import fields
            "vordergrunddienst_cells": current_layout.get("vordergrunddienst_cells", {}),
            "hintergrunddienst_cells": current_layout.get("hintergrunddienst_cells", {}),
            "date_cells": current_layout.get("date_cells", {}),
            "weekday_date_cells": current_layout.get("weekday_date_cells", {}),
            "feiertage": current_layout.get("feiertage", []),
        }
        save_layout(new_layout)
        st.success("Layout gespeichert und neu geladen.")
        st.rerun()

# ===========================================================================
# TAB 6 — Rapporte-Pools (BH + LI)
# ===========================================================================

elif page == "📊 Rapporte-Pools":
    st.subheader("Rapporte-Pools")
    st.caption(
        "Die Pools werden in der definierten Reihenfolge durchlaufen, bis eine "
        "verfügbare Person gefunden wird."
    )

    pools_data = load_meeting_pools()
    all_staff_names = sorted(sched.staff_by_name.keys())

    def _render_pool_editor(meeting_key: str, cfg: dict, default_site: str) -> None:
        """Render all pools + settings for one rapport inside an active tab."""
        prefix = (meeting_key
                  .replace("|", "_").replace(" ", "_").replace(":", "")
                  .replace("/", "_").replace("(", "").replace(")", ""))
        pools = cfg.get("pools", [])

        for i, pool in enumerate(pools):
            with st.container(border=True):
                header_col, up_col, down_col = st.columns([8, 1, 1])
                header_col.markdown(
                    f"<span style='font-size:1.05rem; font-weight:700; color:#4A90D9;'>"
                    f"Pool {i+1}</span>", unsafe_allow_html=True)
                if i > 0:
                    if up_col.button("↑", key=f"{prefix}_p{i}_up"):
                        pools[i], pools[i-1] = pools[i-1], pools[i]
                        cfg["pools"] = pools
                        pools_data[meeting_key] = cfg
                        save_meeting_pools(pools_data); st.rerun()
                if i < len(pools) - 1:
                    if down_col.button("↓", key=f"{prefix}_p{i}_down"):
                        pools[i], pools[i+1] = pools[i+1], pools[i]
                        cfg["pools"] = pools
                        pools_data[meeting_key] = cfg
                        save_meeting_pools(pools_data); st.rerun()

                pc1, pc2, pc3 = st.columns(3)
                current_type = pool.get("type", "names")
                type_display_idx = _POOL_TYPES.index(current_type) if current_type in _POOL_TYPES else 0
                pool_type_display = pc1.selectbox("Typ", options=_POOL_TYPES_DISPLAY,
                    index=type_display_idx, key=f"{prefix}_p{i}_type")
                pool_type = _POOL_TYPES[_POOL_TYPES_DISPLAY.index(pool_type_display)]
                pool["type"] = pool_type
                pool_site = pc2.selectbox("Standort", options=_SITE_OPTIONS,
                    index=_SITE_OPTIONS.index(pool.get("site", cfg.get("site", default_site))),
                    key=f"{prefix}_p{i}_site")
                pool["site"] = pool_site
                roter_text = pc3.checkbox("Roter Text",
                    value=(pool.get("style") == "red_bold"), key=f"{prefix}_p{i}_rot")
                pool["style"] = "red_bold" if roter_text else None

                if pool_type == "names":
                    current_names = pool.get("names") or []
                    selected_names = st.multiselect("Namen", options=all_staff_names,
                        default=[n for n in current_names if n in all_staff_names],
                        key=f"{prefix}_p{i}_names")
                    pool["names"] = selected_names if selected_names else []
                if pool_type == "group":
                    current_group = pool.get("group", "AA")
                    group_display_idx = _GROUP_OPTIONS.index(current_group) if current_group in _GROUP_OPTIONS else 0
                    pool_group_display = st.selectbox("Gruppe", options=_GROUP_DISPLAY,
                        index=group_display_idx, key=f"{prefix}_p{i}_group")
                    pool["group"] = _GROUP_OPTIONS[_GROUP_DISPLAY.index(pool_group_display)]
                if pool_type == "hintergrund_vortag":
                    st.caption("Person wird automatisch aus dem Hintergrund-Dienst des Vortags bestimmt.")

                is_auto_type = pool_type in ("spaetdienst_aa", "hintergrund_vortag")
                cb_col1, cb_col2 = st.columns(2)
                excl_spaet = cb_col1.checkbox("Spätdienst ausschließen",
                    value=bool(pool.get("exclude_spaetdienst")),
                    key=f"{prefix}_p{i}_excl_spaet", disabled=is_auto_type)
                pool["exclude_spaetdienst"] = pool_site if excl_spaet else None
                excl_hintergrund = cb_col2.checkbox("Hintergrund ausschließen",
                    value=bool(pool.get("exclude_hintergrund")),
                    key=f"{prefix}_p{i}_excl_hintergr", disabled=is_auto_type)
                pool["exclude_hintergrund"] = excl_hintergrund

                current_excluded = pool.get("exclude_names") or []
                excluded_names = st.multiselect("Ausgeschlossene Personen",
                    options=all_staff_names,
                    default=[n for n in current_excluded if n in all_staff_names],
                    key=f"{prefix}_p{i}_excl_names")
                pool["exclude_names"] = excluded_names if excluded_names else None

                eid_str = st.text_input("Ausschluss pro Tag",
                    value=_exclude_if_day_to_str(pool.get("exclude_if_day")),
                    key=f"{prefix}_p{i}_eid",
                    help="Format: 'Donnerstag: Name1, Name2; Freitag: Name3'")
                pool["exclude_if_day"] = _str_to_exclude_if_day(eid_str)

                if i == len(pools) - 1 and len(pools) > 1:
                    if st.button("Pool entfernen", key=f"{prefix}_p{i}_remove"):
                        pools.pop(i)
                        cfg["pools"] = pools
                        save_meeting_pools(pools_data); st.rerun()

        if st.button("Pool hinzufügen", key=f"{prefix}_add_pool"):
            pools.append({"type": "names", "names": [], "site": cfg.get("site", default_site)})
            cfg["pools"] = pools
            save_meeting_pools(pools_data); st.rerun()

        st.markdown("---")
        st.markdown("**Fallback-Einstellungen**")
        c1, c2 = st.columns(2)
        cfg["fallback_text"] = c1.text_input("Fallback-Text",
            value=cfg.get("fallback_text", "FÄLLT AUS"), key=f"{prefix}_fb_text")
        cfg["roter_fallback_text"] = c2.checkbox("Roter Text",
            value=cfg.get("roter_fallback_text", True), key=f"{prefix}_fb_rot")

        st.markdown("---")
        st.markdown("**Statistik-Einstellungen**")
        stat_col1, _ = st.columns([1, 3])
        statistik_führen = stat_col1.checkbox(
            "Statistik führen",
            value=cfg.get("statistik_führen", False),
            key=f"{prefix}_statistik",
            help="Wenn aktiviert, wird die Zuteilung dieses Rapports über Wochen hinweg verfolgt und ausgeglichen.",
        )
        cfg["statistik_führen"] = statistik_führen

        if statistik_führen:
            st.caption("Gewichtung pro Person (Standard 1.0). Niedrigere Werte bedeuten seltenere Zuteilung im Verhältnis zu anderen.")
            _wt_name_set = set()
            for pool in cfg.get("pools", []):
                if pool.get("type") == "names":
                    _wt_name_set.update(pool.get("names") or [])
                elif pool.get("type") == "group" and pool.get("group"):
                    try:
                        _wt_name_set.update(
                            sched._group_names(pool["group"], pool.get("site", cfg.get("site", "")))
                        )
                    except Exception:
                        pass
            all_pool_names = sorted(_wt_name_set)
            current_weights = cfg.get("stats_weight", {})
            new_weights = {}
            if all_pool_names:
                sw_cols = st.columns(min(len(all_pool_names), 4))
                for wi, name in enumerate(all_pool_names):
                    with sw_cols[wi % 4]:
                        new_weights[name] = st.number_input(
                            name, min_value=0.1, max_value=10.0,
                            value=float(current_weights.get(name, 1.0)),
                            step=0.1, format="%.1f",
                            key=f"{prefix}_sw_{name}",
                        )
            else:
                st.caption("Keine Personen in 'Namen'-Pools gefunden.")
            cfg["stats_weight"] = new_weights
        else:
            cfg.pop("stats_weight", None)

        cfg["pools"] = pools
        pools_data[meeting_key] = cfg

        st.markdown("---")
        if st.button("Änderungen speichern", type="primary", key=f"{prefix}_save"):
            validation_errors = []
            for pool in cfg.get("pools", []):
                if pool.get("type") == "names" and not pool.get("names"):
                    validation_errors.append(f"Pool ohne Namen.")
                if pool.get("type") == "group" and not pool.get("group"):
                    validation_errors.append(f"Pool ohne Gruppe.")
            if validation_errors:
                for err in validation_errors: st.error(f"• {err}")
            else:
                for pool in cfg.get("pools", []):
                    for k in list(pool.keys()):
                        if pool[k] is None or pool[k] == "" or pool[k] == [] or pool[k] is False:
                            if k not in ("type", "names", "group", "site", "exclude_hintergrund"):
                                del pool[k]
                save_meeting_pools(pools_data)
                st.success("Gespeichert.")
                st.rerun()

    # --- Outer BH / LI tabs ---
    site_tab_bh, site_tab_li = st.tabs(["📊 Rapporte-Pools BH", "📊 Rapporte-Pools LI"])

    with site_tab_bh:
        bh_pools = {k: v for k, v in pools_data.items() if v.get("site") == "BH"}
        if bh_pools:
            bh_tab_labels = [k.split("|", 1)[-1].strip() for k in bh_pools.keys()]
            bh_tabs = st.tabs(bh_tab_labels)
            for tab, (meeting_key, cfg) in zip(bh_tabs, bh_pools.items()):
                with tab:
                    _render_pool_editor(meeting_key, cfg, "BH")
        else:
            st.caption("Keine BH-Rapporte vorhanden.")

    with site_tab_li:
        li_pools = {k: v for k, v in pools_data.items() if v.get("site") == "LI"}
        if li_pools:
            li_tab_labels = [k.split("|", 1)[-1].strip() for k in li_pools.keys()]
            li_tabs = st.tabs(li_tab_labels)
            for tab, (meeting_key, cfg) in zip(li_tabs, li_pools.items()):
                with tab:
                    _render_pool_editor(meeting_key, cfg, "LI")
        else:
            st.caption("Keine LI-Rapporte vorhanden.")


# ===========================================================================
# TAB 7 — Frontarzt
# ===========================================================================

elif page == "👨‍⚕️ Frontarzt":
    st.subheader("Frontarzt-Einstellungen")
    st.caption("Ausschlussregeln und standortspezifische Einstellungen für Frontarztdienste.")

    _fr_rules_path = Path(sched._staff_json).parent / "fr_rules.json"

    def _load_fr_rules_ui() -> dict:
        if _fr_rules_path.exists():
            with open(_fr_rules_path, encoding="utf-8") as f:
                return json.load(f)
        return {"exclude_names": [], "exclude_if_day": {}, "exclude_from_frontarzt": []}

    def _save_fr_rules_ui(data: dict) -> None:
        with open(_fr_rules_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        sched.reload_fr_rules()

    fr_rules = _load_fr_rules_ui()
    all_staff_names = sorted(sched.staff_by_name.keys())
    og_list = sched.OG_LIST

    # ---- Standort-spezifische Regeln (moved from Personalverwaltung) ----
    st.markdown("### Standort-spezifische Regeln")
    site_rules = sched.SITE_RULES if sched.SITE_RULES else {
        "BH": {"no_oa_vormittag": False},
        "LI": {"no_oa_vormittag": True}
    }
    src1, src2 = st.columns(2)
    bh_no_oa = src1.checkbox(
        "BH: Kein OA im Frontarzt-Vormittagsdienst",
        value=site_rules.get("BH", {}).get("no_oa_vormittag", False),
        key="bh_no_oa_vormittag",
        help="Wenn aktiviert: erste FR-Zelle BH nur mit LA besetzt."
    )
    li_no_oa = src2.checkbox(
        "LI: Kein OA im Frontarzt-Vormittagsdienst",
        value=site_rules.get("LI", {}).get("no_oa_vormittag", True),
        key="li_no_oa_vormittag",
        help="Wenn aktiviert: erste FR-Zelle LI nur mit LA besetzt."
    )

    st.markdown("**Ausschluss Vormittag / Nachmittag**")
    st.caption("Format: 'Vormittag: Name1, Name2; Nachmittag: Name3' — gilt für BH und LI.")

    def _fr_slot_excl_to_str(vormittag: list, nachmittag: list) -> str:
        parts = []
        if vormittag:
            parts.append("Vormittag: " + ", ".join(vormittag))
        if nachmittag:
            parts.append("Nachmittag: " + ", ".join(nachmittag))
        return "; ".join(parts)

    def _str_to_fr_slot_excl(s: str):
        vormittag, nachmittag = [], []
        for part in s.split(";"):
            part = part.strip()
            if part.lower().startswith("vormittag:"):
                names = [n.strip() for n in part[len("vormittag:"):].split(",") if n.strip()]
                vormittag.extend(names)
            elif part.lower().startswith("nachmittag:"):
                names = [n.strip() for n in part[len("nachmittag:"):].split(",") if n.strip()]
                nachmittag.extend(names)
        return vormittag, nachmittag

    _bh_vorm = site_rules.get("BH", {}).get("fr_excluded_vormittag", [])
    _bh_nach = site_rules.get("BH", {}).get("fr_excluded_nachmittag", [])
    fr_slot_excl_str = st.text_input(
        "Ausschluss Vormittag / Nachmittag",
        value=_fr_slot_excl_to_str(_bh_vorm, _bh_nach),
        key="fr_slot_excl",
        help="Format: 'Vormittag: Name1, Name2; Nachmittag: Name3'",
        label_visibility="collapsed",
    )
    _fr_excl_vorm, _fr_excl_nach = _str_to_fr_slot_excl(fr_slot_excl_str)

    st.divider()

    # ---- Ausgeschlossene Personen ----
    st.markdown("### Ausgeschlossene Personen")
    excl_names = st.multiselect(
        "Personen die nie als Frontarzt eingeteilt werden",
        options=all_staff_names,
        default=[n for n in fr_rules.get("exclude_names", []) if n in all_staff_names],
        key="fr_excl_names",
    )

    # ---- Ausschluss pro Tag ----
    st.markdown("### Ausschluss pro Tag")
    eid_str = st.text_input(
        "Ausschluss pro Tag",
        value=_exclude_if_day_to_str(fr_rules.get("exclude_if_day")),
        key="fr_excl_per_day",
        help="Format: 'Donnerstag: Name1, Name2; Freitag: Name3'",
    )

    st.divider()

    # ---- Von Frontarzt ausschließen (OG-based) ----
    st.markdown("### Von Frontarzt ausschließen")
    st.caption("OGs deren Mitglieder nie als Frontarzt eingeteilt werden")
    excl_og = st.multiselect(
        "OGs auswählen",
        options=og_list,
        default=[o for o in fr_rules.get("exclude_from_frontarzt", []) if o in og_list],
        key="fr_excl_og",
    )

    st.divider()

    if st.button("Frontarzt-Einstellungen speichern", type="primary", key="save_fr_rules_btn"):
        # Save site rules into staff.json
        new_site_rules = {
            "BH": {
                "no_oa_vormittag": bh_no_oa,
                "fr_excluded_vormittag": _fr_excl_vorm,
                "fr_excluded_nachmittag": _fr_excl_nach,
            },
            "LI": {
                "no_oa_vormittag": li_no_oa,
                "fr_excluded_vormittag": _fr_excl_vorm,
                "fr_excluded_nachmittag": _fr_excl_nach,
            },
        }
        sched.SITE_RULES.update(new_site_rules)
        # Read current staff.json and update site_rules
        _staff_path = Path(sched._staff_json)
        with open(_staff_path, encoding="utf-8") as f:
            staff_json = json.load(f)
        staff_json["site_rules"] = new_site_rules
        with open(_staff_path, "w", encoding="utf-8") as f:
            json.dump(staff_json, f, ensure_ascii=False, indent=2)

        # Save fr_rules.json
        _save_fr_rules_ui({
            "exclude_names": excl_names,
            "exclude_if_day": _str_to_exclude_if_day(eid_str) or {},
            "exclude_from_frontarzt": excl_og,
        })
        st.success("Frontarzt-Einstellungen gespeichert.")
        st.rerun()


# ===========================================================================
# Organgruppen parent placeholder
# ===========================================================================

# ===========================================================================
# Organgruppen Verwalten
# ===========================================================================

elif page == "🏥 Organgruppen Verwalten":
    st.subheader("Organgruppen verwalten")
    st.caption("Hinzufügen oder Entfernen von Organgruppen.")
    
    # Load current OGs
    og_file = Path(sched._staff_json).parent / "organgruppen.json"
    if og_file.exists():
        with open(og_file, encoding="utf-8") as f:
            og_data = json.load(f)
            current_ogs = og_data.get("organgruppen", [])
    else:
        current_ogs = ["MSK", "Neuro", "Onko", "Thorax", "Abdomen", "Mammo", "Intervention/ Vaskulär", "Nuklearmedizin"]
    
    # Display current OGs
    st.markdown("### Aktuelle Organgruppen")
    for og in current_ogs:
        col1, col2 = st.columns([4, 1])
        with col1:
            st.markdown(f"**{og}**")
        with col2:
            if st.button("Löschen", key=f"delete_og_{og}"):
                # Confirm deletion
                st.session_state[f"confirm_delete_{og}"] = True
        
        # Show confirmation dialog
        if st.session_state.get(f"confirm_delete_{og}", False):
            st.warning(f"⚠️ Möchten Sie '{og}' wirklich löschen? Dies entfernt die OG aus allen Konfigurationen und Personal-Rotationen.")
            col_yes, col_no = st.columns(2)
            with col_yes:
                if st.button("Ja, löschen", key=f"confirm_yes_{og}", type="primary"):
                    # Remove from organgruppen.json
                    current_ogs.remove(og)
                    with open(og_file, "w", encoding="utf-8") as f:
                        json.dump({"organgruppen": current_ogs}, f, ensure_ascii=False, indent=2)
                    
                    # Remove from og_rules.json
                    og_rules = load_og_rules()
                    if og in og_rules.get("og_priority_order", []):
                        og_rules["og_priority_order"].remove(og)
                    for key in ["og_weights_oa", "og_weights_aa", "og_max_fas", "og_max_aas"]:
                        if og in og_rules.get(key, {}):
                            del og_rules[key][og]
                    for key in ["rotation_or_leader_only", "warn_kein_aa", "warn_weniger_als_2fa", "warn_kein_fa_site"]:
                        if og in og_rules.get(key, []):
                            og_rules[key].remove(og)
                    save_og_rules(og_rules)
                    
                    # Remove from staff.json (rotations, leads_ogs, avoid_ogs)
                    for staff in sched.staff_by_name.values():
                        if og in staff.rotations:
                            staff.rotations.remove(og)
                        if og in staff.leads_ogs:
                            staff.leads_ogs.remove(og)
                        if og in staff.avoid_ogs:
                            staff.avoid_ogs.remove(og)
                    save_staff_to_json()
                    
                    # Reload
                    sched.reload_og_rules()
                    
                    del st.session_state[f"confirm_delete_{og}"]
                    st.success(f"'{og}' wurde gelöscht.")
                    st.rerun()
            
            with col_no:
                if st.button("Abbrechen", key=f"confirm_no_{og}"):
                    del st.session_state[f"confirm_delete_{og}"]
                    st.rerun()
    
    # Add new OG
    st.divider()
    st.markdown("### Neue Organgruppe hinzufügen")
    
    with st.form("add_og_form"):
        new_og_name = st.text_input("Name der neuen Organgruppe")
        submitted = st.form_submit_button("Hinzufügen", type="primary")
        
        if submitted:
            if not new_og_name.strip():
                st.error("Bitte geben Sie einen Namen ein.")
            elif new_og_name.strip() in current_ogs:
                st.error(f"'{new_og_name.strip()}' existiert bereits.")
            else:
                # Add to organgruppen.json
                current_ogs.append(new_og_name.strip())
                with open(og_file, "w", encoding="utf-8") as f:
                    json.dump({"organgruppen": current_ogs}, f, ensure_ascii=False, indent=2)
                
                # Add to og_rules.json with defaults
                og_rules = load_og_rules()
                og_rules["og_priority_order"].append(new_og_name.strip())
                og_rules.setdefault("og_weights_oa", {})[new_og_name.strip()] = 0.6
                og_rules.setdefault("og_weights_aa", {})[new_og_name.strip()] = 0.6
                og_rules.setdefault("og_max_fas", {})[new_og_name.strip()] = None
                og_rules.setdefault("og_max_aas", {})[new_og_name.strip()] = None
                save_og_rules(og_rules)
                
                # Reload
                sched.reload_og_rules()
                
                st.success(f"'{new_og_name.strip()}' wurde hinzugefügt.")
                st.rerun()


# ===========================================================================
# TAB 5.6 — Organgruppen Regeln
# ===========================================================================

elif page == "🏥 Organgruppen Regeln":
    
    og_rules = load_og_rules()
    
    # Section 1: Priority Order
    st.markdown("### Organgruppen-Priorität")
    st.caption("Reihenfolge bei gleicher Auslastung (oben = höchste Priorität). Diese Reihenfolge gilt nur wenn 'Zufällige Zuteilung' deaktiviert ist.")
    
    current_order = og_rules.get("og_priority_order", sched.OG_LIST)
    
    # Ensure all OGs are in the list
    for og in sched.OG_LIST:
        if og not in current_order:
            current_order.append(og)
    
    # Display with move buttons
    changed = False
    for i, og in enumerate(current_order):
        col1, col2, col3 = st.columns([6, 1, 1])
        col1.write(f"**{i+1}.** {og}")
        
        if i > 0:
            if col2.button("↑", key=f"og_up_{i}"):
                current_order[i], current_order[i-1] = current_order[i-1], current_order[i]
                og_rules["og_priority_order"] = current_order
                save_og_rules(og_rules)
                st.rerun()
        
        if i < len(current_order) - 1:
            if col3.button("↓", key=f"og_down_{i}"):
                current_order[i], current_order[i+1] = current_order[i+1], current_order[i]
                og_rules["og_priority_order"] = current_order
                save_og_rules(og_rules)
                st.rerun()
    
    use_random = st.checkbox(
        "Zufällige Zuteilung bei gleicher Auslastung",
        value=og_rules.get("use_random_og_selection", False),
        help="Wenn aktiviert, wird bei gleicher Auslastung mehrerer OGs zufällig gewählt. Wenn deaktiviert, wird die obige Prioritäts-Reihenfolge verwendet."
    )
    
    if st.button("Priorität speichern", type="primary", key="save_og_priority"):
        og_rules["use_random_og_selection"] = use_random
        save_og_rules(og_rules)
        st.success("Organgruppen-Priorität gespeichert!")
        st.rerun()
    
    st.divider()
    
    # Section 1b: OG Weights for OAs
    st.markdown("### OG-Gewichtung für OAs (Oberärzte)")
    st.caption("Gewichtung zwischen 0.1 und 1.0. Niedrigere Werte erlauben mehr Zuweisungen pro Person. "
              "Beispiel: 0.4 ermöglicht 2 Zuweisungen (0.4 + 0.4 = 0.8 ≤ 1.0), 0.6 ermöglicht max. 1 Zuweisung.")
    
    og_weights_oa = og_rules.get("og_weights_oa", og_rules.get("og_weights", {}))
    
    # Display in 2 columns
    col1, col2 = st.columns(2)
    updated_weights_oa = {}
    
    for i, og in enumerate(sched.OG_LIST):
        col = col1 if i % 2 == 0 else col2
        
        default = 0.4 if og in ["Mammo", "Intervention/ Vaskulär"] else 0.6
        with col:
            weight = st.number_input(
                og,
                min_value=0.1,
                max_value=1.0,
                value=og_weights_oa.get(og, default),
                step=0.1,
                format="%.1f",
                key=f"og_weight_oa_{og}"
            )
            updated_weights_oa[og] = weight
    
    if st.button("OG-Gewichtungen für OAs speichern", type="primary", key="save_og_weights_oa"):
        og_rules["og_weights_oa"] = updated_weights_oa
        save_og_rules(og_rules)
        st.success("OG-Gewichtungen für OAs gespeichert!")
        st.rerun()
    
    st.divider()
    
    # Section 1c: OG Weights for AAs
    st.markdown("### OG-Gewichtung für AAs (Assistenzärzte)")
    st.caption("Gewichtung zwischen 0.1 und 1.0. Niedrigere Werte erlauben mehr Zuweisungen pro Person. "
              "Beispiel: 0.4 ermöglicht 2 Zuweisungen (0.4 + 0.4 = 0.8 ≤ 1.0), 0.6 ermöglicht max. 1 Zuweisung.")
    
    og_weights_aa = og_rules.get("og_weights_aa", og_rules.get("og_weights", {}))
    
    # Display in 2 columns
    col1, col2 = st.columns(2)
    updated_weights_aa = {}
    
    for i, og in enumerate(sched.OG_LIST):
        col = col1 if i % 2 == 0 else col2
        
        default = 0.4 if og in ["Mammo", "Intervention/ Vaskulär"] else 0.6
        with col:
            weight = st.number_input(
                og,
                min_value=0.1,
                max_value=1.0,
                value=og_weights_aa.get(og, default),
                step=0.1,
                format="%.1f",
                key=f"og_weight_aa_{og}"
            )
            updated_weights_aa[og] = weight
    
    if st.button("OG-Gewichtungen für AAs speichern", type="primary", key="save_og_weights_aa"):
        og_rules["og_weights_aa"] = updated_weights_aa
        save_og_rules(og_rules)
        st.success("OG-Gewichtungen für AAs gespeichert!")
        st.rerun()
    
    st.divider()
    
    # Section 1d: Max FAs and Max AAs
    st.markdown("### Maximum FAs und AAs pro OG")
    st.caption("Legt fest, wie viele FAs/AAs maximal pro Tag in eine OG zugewiesen werden. "
              "0 = kein Limit. Verhindert, dass niedrig-gewichtete OGs (0.4) immer vollständig gefüllt werden.")
    
    og_max_fas = og_rules.get("og_max_fas", {})
    og_max_aas = og_rules.get("og_max_aas", {})
    
    # Display in 3 columns: OG name, Max FAs, Max AAs
    st.markdown("**OG | Max FAs | Max AAs**")
    updated_max_fas = {}
    updated_max_aas = {}
    
    for og in sched.OG_LIST:
        col1, col2, col3 = st.columns([2, 1, 1])
        
        with col1:
            st.markdown(f"**{og}**")
        
        with col2:
            max_fa = st.number_input(
                "Max FAs",
                min_value=0,
                max_value=10,
                value=og_max_fas.get(og) if og_max_fas.get(og) is not None else 0,
                step=1,
                key=f"og_max_fa_{og}",
                label_visibility="collapsed"
            )
            updated_max_fas[og] = max_fa if max_fa > 0 else None
        
        with col3:
            max_aa = st.number_input(
                "Max AAs",
                min_value=0,
                max_value=10,
                value=og_max_aas.get(og) if og_max_aas.get(og) is not None else 0,
                step=1,
                key=f"og_max_aa_{og}",
                label_visibility="collapsed"
            )
            updated_max_aas[og] = max_aa if max_aa > 0 else None
    
    if st.button("Max FAs/AAs speichern", type="primary", key="save_og_maxs"):
        og_rules["og_max_fas"] = updated_max_fas
        og_rules["og_max_aas"] = updated_max_aas
        save_og_rules(og_rules)
        st.success("Maximum FAs/AAs gespeichert!")
        st.rerun()
    
    st.divider()
    
    # Section 2: Special Rules (moved from Layout Editor)
    st.markdown("### Organgruppen-Sonderregeln")
    st.caption("Regeln für automatische Warnungen und spezielle Zuweisungen")

    site_cov_mode = st.radio(
        "Priorisierung bei OGs mit Pflicht-Abdeckung beider Standorte",
        options=[
            "Standortabdeckung vor Avoid/Rotation",
            "Avoid/Rotation vor Standortabdeckung",
        ],
        index=0 if og_rules.get("site_coverage_over_avoid", True) else 1,
        key="site_coverage_mode",
        help=(
            "Gilt nur für OGs mit Pflicht-Abdeckung beider Standorte, wenn nur ein "
            "Standort besetzt ist. 'Standortabdeckung vor Avoid/Rotation': eine Person "
            "vom fehlenden Standort wird auch dann gewählt, wenn sie diese OG meidet. "
            "'Avoid/Rotation vor Standortabdeckung': die Meiden-/Rotations-Präferenz hat "
            "Vorrang (bisheriges Verhalten). Die Rotations-Reihenfolge innerhalb einer "
            "Gruppe bleibt in beiden Fällen gleich."
        ),
    )

    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("**Nur Rotation/Leader**")
        st.caption("OGs die nur an Personen mit Rotation oder Leader zugeteilt werden")
        rotation_only = st.multiselect(
            "Organgruppen",
            options=sched.OG_LIST,
            default=sorted([og for og in og_rules.get("rotation_or_leader_only", []) if og != "Laufen"]),
            key="rotation_only_select",
            label_visibility="collapsed"
        )
        
        st.markdown("**WENIGER ALS 2FA Warnung**")
        st.caption("OGs die 'WENIGER ALS 2FA' anzeigen wenn < 2 Fachärzte")
        warn_2fa = st.multiselect(
            "Organgruppen",
            options=sched.OG_LIST,
            default=sorted(og_rules.get("warn_weniger_als_2fa", [])),
            key="warn_2fa_select",
            label_visibility="collapsed"
        )
    
    with col2:
        st.markdown("**KEIN AA Warnung**")
        st.caption("OGs die 'KEIN AA' anzeigen wenn kein Assistenzarzt zugeteilt")
        warn_aa = st.multiselect(
            "Organgruppen",
            options=sched.OG_LIST,
            default=sorted(og_rules.get("warn_kein_aa", [])),
            key="warn_aa_select",
            label_visibility="collapsed"
        )
        
        st.markdown("**OGs mit Pflicht-Abdeckung beider Standorte**")
        st.caption("Für diese OGs versucht die Zuteilung, an beiden Standorten (BH und LI) einen FA zu platzieren. Gelingt dies nicht, wird 'KEIN FA IN BH/LI' als Warnung eingetragen.")
        warn_site = st.multiselect(
            "Organgruppen",
            options=sched.OG_LIST,
            default=sorted(og_rules.get("warn_kein_fa_site", [])),
            key="warn_site_select",
            label_visibility="collapsed"
        )
        
        st.markdown("**Von Rapporten ausschließen**")
        st.caption("OGs die nie in Rapporte-Pools zugewiesen werden")
        exclude_rapporte = st.multiselect(
            "Organgruppen",
            options=sched.OG_LIST,
            default=sorted(og_rules.get("exclude_from_rapporte", [])),
            key="exclude_rapporte_select",
            label_visibility="collapsed"
        )

        st.markdown("**US-Vertretung Pools**")
        st.caption("Für OGs mit 'KEIN FA IN SITE' Warnung: Prioritätsliste der FAs für US-Vertretung (kommagetrennt). Nur FAs vom fehlenden Standort werden berücksichtigt.")
        us_vertretung_pools = dict(og_rules.get("us_vertretung_pools", {}))
        for og in warn_site:
            current_list = ", ".join(us_vertretung_pools.get(og, []))
            new_str = st.text_input(
                f"US-Vertretung Pool: {og}",
                value=current_list,
                key=f"us_pool_{og}",
                help="Format: 'Name1, Name2, Name3' — in Prioritätsreihenfolge"
            )
            us_vertretung_pools[og] = [n.strip() for n in new_str.split(",") if n.strip()]

        st.markdown("**Organgruppen Vertretung**")
        st.caption("OGs die eine Vertretung '(Name)' erhalten wenn komplett leer")
        og_vertretung_ogs = st.multiselect(
            "Organgruppen",
            options=sched.OG_LIST,
            default=sorted(og_rules.get("og_vertretung_ogs", [])),
            key="og_vertretung_ogs_select",
            label_visibility="collapsed"
        )
        og_vertretung_pools = dict(og_rules.get("og_vertretung_pools", {}))
        for og in og_vertretung_ogs:
            current_list = ", ".join(og_vertretung_pools.get(og, []))
            new_str = st.text_input(
                f"Vertretung Pool: {og}",
                value=current_list,
                key=f"og_vert_pool_{og}",
                help="Format: 'Name1, Name2, Name3' — in Prioritätsreihenfolge"
            )
            og_vertretung_pools[og] = [n.strip() for n in new_str.split(",") if n.strip()]

    if st.button("Sonderregeln speichern", type="primary", key="save_og_rules"):
        og_rules["rotation_or_leader_only"] = rotation_only
        og_rules["warn_kein_aa"] = warn_aa
        og_rules["warn_weniger_als_2fa"] = warn_2fa
        og_rules["warn_kein_fa_site"] = warn_site
        og_rules["exclude_from_rapporte"] = exclude_rapporte
        og_rules["us_vertretung_pools"] = us_vertretung_pools
        og_rules["og_vertretung_ogs"] = og_vertretung_ogs
        og_rules["og_vertretung_pools"] = og_vertretung_pools
        og_rules["site_coverage_over_avoid"] = (
            site_cov_mode == "Standortabdeckung vor Avoid/Rotation"
        )
        save_og_rules(og_rules)
        st.success("Organgruppen-Sonderregeln gespeichert!")
        st.rerun()


# ===========================================================================
# TAB 6 — Radiologe in Laufen
# ===========================================================================




# ===========================================================================
# PAGE — Rapport-Statistik
# ===========================================================================

elif page == "📈 Rapport-Statistik":
    st.subheader("Rapport-Statistik")
    st.caption(
        "Zuteilungs-Verlauf für Rapporte mit aktivierter Statistik. "
        "Einträge können manuell korrigiert werden."
    )

    STATS_JSON_PATH = Path(__file__).parent / "stats.json"

    def _load_stats_ui() -> dict:
        if not STATS_JSON_PATH.exists():
            return {}
        with open(STATS_JSON_PATH, encoding="utf-8") as f:
            return json.load(f)

    def _save_stats_ui(data: dict) -> None:
        tmp = str(STATS_JSON_PATH) + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        os.replace(tmp, str(STATS_JSON_PATH))

    pools_data = load_meeting_pools()
    tracked = {k: v for k, v in pools_data.items() if v.get("statistik_führen")}

    if not tracked:
        st.info("Keine Rapporte mit aktivierter Statistik gefunden. "
                "Statistik kann in den Rapporte-Pools unter 'Statistik-Einstellungen' aktiviert werden.")
        st.stop()

    stats = _load_stats_ui()

    # One tab per tracked rapport
    tab_labels = list(tracked.keys())
    rapport_tabs = st.tabs(tab_labels)

    for tab, (meeting_key, cfg) in zip(rapport_tabs, tracked.items()):
      with tab:
        rapport_stats = stats.get(meeting_key, {})
        stats_weight  = cfg.get("stats_weight", {})

        # Collect all names from pools — expand both "names" and "group" types
        _pool_name_set = set()
        for pool in cfg.get("pools", []):
            if pool.get("type") == "names":
                _pool_name_set.update(pool.get("names") or [])
            elif pool.get("type") == "group" and pool.get("group"):
                try:
                    _pool_name_set.update(
                        sched._group_names(pool["group"], pool.get("site", cfg.get("site", "")))
                    )
                except Exception:
                    pass
        pool_names = sorted(
            n for n in _pool_name_set
            if rapport_stats.get(n, {}).get("count", 0) > 0
        )

        if not pool_names:
            st.caption("Noch keine Einträge vorhanden.")
            continue

        # Compute max ratio for parity calculation
        def _ratio(name):
            e = rapport_stats.get(name, {})
            c = e.get("count", 0)
            w = stats_weight.get(name, 1.0)
            return c / w if w else 0.0

        max_ratio = max((_ratio(n) for n in pool_names), default=0.0)

        import math
        # Build summary table
        rows = []
        for name in pool_names:
            entry  = rapport_stats.get(name, {"count": 0, "history": []})
            count  = entry.get("count", 0)
            weight = stats_weight.get(name, 1.0)
            ratio  = _ratio(name)
            # Rapporte needed to reach parity with the most-assigned person
            needed = max(0, math.ceil(max_ratio * weight - count))
            rows.append({
                "Name":                              name,
                "Gewichtung":                        weight,
                "Anzahl":                            count,
                "Ausstehend bis Parität":            needed,
            })

        df_stats = pd.DataFrame(rows)
        st.dataframe(df_stats, use_container_width=True, hide_index=True)

        # Build chronological history table across all names
        history_rows = []
        for name in pool_names:
            entry = rapport_stats.get(name, {"count": 0, "history": []})
            for kw in entry.get("history", []):
                history_rows.append({"KW": kw, "Person": name})

        if history_rows:
            history_rows.sort(key=lambda r: r["KW"])
            df_history = pd.DataFrame(history_rows)
            st.markdown("**Verlauf (chronologisch)**")
            st.dataframe(df_history, use_container_width=True, hide_index=True)
        else:
            st.caption("Noch keine Einträge vorhanden.")

        # Manual correction expander
        with st.expander(f"Einträge bearbeiten"):
            edit_name = st.selectbox(
                "Person auswählen",
                options=pool_names,
                key=f"stat_edit_name_{meeting_key}",
            )
            if edit_name:
                current_entry  = rapport_stats.get(edit_name, {"count": 0, "history": []})
                current_history = current_entry.get("history", [])

                st.markdown(f"**Aktueller Verlauf für {edit_name}:** "
                            f"{', '.join(current_history) if current_history else '—'}")

                ec1, ec2 = st.columns(2)

                # Add KW entry — year selector + KW dropdown
                import datetime as _dt
                _current_year = _dt.date.today().year
                add_year = ec1.number_input(
                    "Jahr",
                    min_value=2020,
                    max_value=_current_year + 1,
                    value=_current_year,
                    step=1,
                    key=f"stat_add_year_{meeting_key}_{edit_name}",
                )
                kw_options = [f"{int(add_year)}-KW{k:02d}" for k in range(1, 54)]
                add_kw = ec1.selectbox(
                    "KW auswählen",
                    options=kw_options,
                    key=f"stat_add_kw_{meeting_key}_{edit_name}",
                )
                if ec1.button("Hinzufügen", key=f"stat_add_btn_{meeting_key}_{edit_name}"):
                    entry = stats.setdefault(meeting_key, {}).setdefault(
                        edit_name, {"count": 0, "history": []}
                    )
                    entry["history"].append(add_kw)
                    entry["count"] = len(entry["history"])
                    _save_stats_ui(stats)
                    st.success(f"{add_kw} für {edit_name} hinzugefügt.")
                    st.rerun()

                # Remove KW entry
                if current_history:
                    remove_kw = ec2.selectbox(
                        "KW entfernen",
                        options=current_history,
                        key=f"stat_remove_kw_{meeting_key}_{edit_name}",
                    )
                    if ec2.button("Entfernen", key=f"stat_remove_btn_{meeting_key}_{edit_name}"):
                        entry = stats.setdefault(meeting_key, {}).setdefault(
                            edit_name, {"count": 0, "history": []}
                        )
                        if remove_kw in entry["history"]:
                            entry["history"].remove(remove_kw)
                            entry["count"] = len(entry["history"])
                            _save_stats_ui(stats)
                            st.success(f"{remove_kw} für {edit_name} entfernt.")
                            st.rerun()

        st.divider()
